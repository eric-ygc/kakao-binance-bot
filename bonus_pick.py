"""
보너스픽 — No more → Done/OK 자동 클릭 앱

실행: python bonus_pick.py
설정: bonus_config.json (자동 저장/복원)
"""
import datetime
import json
import queue
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Optional

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

sys.path.insert(0, str(BASE_DIR))

from src.logger_config import setup_logger

try:
    from src.browser_controller import AutoCancelled, click_no_more
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

CONFIG_PATH   = BASE_DIR / "bonus_config.json"
LOG_DATA_PATH = BASE_DIR / "bonus_log_data.json"

DEFAULT_CONFIG = {
    "site_urls":        ["https://dsj44.com/h5/#/login", "", "", "", "", "", "", "", "", ""],
    "chrome_port":      9222,
    "accounts":         [],
    "account_index":    0,
    "schedule_times":   ["", "", "", ""],
    "schedule_enabled": [False, False, False, False],
}

logger = setup_logger("bonus_pick")

# 단계 정의: (표시 이름, 감지 키워드 목록)
STEPS = [
    ("Chrome 실행",  ["Chrome 실행 중"]),
    ("로그인",       ["로그인 페이지 접속", "로그인 필요", "로그인 완료", "세션 유지"]),
    ("홈 이동",      ["홈 접속", "홈으로 이동", "홈 접속 완료"]),
    ("메뉴 선택",    ["Quickly buy", "Invited me"]),
    ("No more",      ["No more 클릭"]),
    ("Done/OK",      ["Done/OK 클릭", "완료"]),
]

# ---------------------------------------------------------------------------
# Neumorphism 색상 팔레트
# ---------------------------------------------------------------------------
NEU_BG    = "#e0e5ec"
NEU_LIGHT = "#ffffff"
NEU_DARK  = "#a3b1c6"

C = {
    "bg":        NEU_BG,
    "panel":     NEU_BG,
    "panel2":    "#d1d9e6",
    "input":     "#d1d9e6",
    "border":    NEU_DARK,
    "fg":        "#31344b",
    "fg_dim":    "#9baacf",
    "fg_bright": "#1a1d2e",
    "accent":    "#5b86e5",
    "yellow":    "#c87800",
    "code_bg":   "#d1d9e6",
    "log_bg":    "#d1d9e6",
    "start":     "#2e7d57",
    "start_hl":  "#3a9b6b",
    "stop":      "#b83232",
    "stop_hl":   "#d43c3c",
    "sel":       "#b8c4d6",
    "error":     "#c0392b",
    "ok":        "#27a06a",
    "system":    "#7f8c9a",
}

# ---------------------------------------------------------------------------
# 설정 로드/저장
# ---------------------------------------------------------------------------

def load_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                data = json.load(f)
            return {**DEFAULT_CONFIG, **data}
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)


def save_config(cfg: dict) -> None:
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"설정 저장 실패: {e}")


# ---------------------------------------------------------------------------
# GUI 앱
# ---------------------------------------------------------------------------

class BonusPickApp(tk.Frame):
    def __init__(self, parent=None, shared_site_url_vars=None) -> None:
        if parent is None:
            self._win = tk.Tk()
            self._win.title("보너스픽")
            self._win.resizable(True, True)
            self._win.minsize(580, 500)
            self._win.configure(bg=C["bg"])
            parent = self._win
        else:
            self._win = None
        super().__init__(parent, bg=C["bg"])
        self._shared_site_url_vars = shared_site_url_vars

        self._auto_cancel_event: threading.Event = threading.Event()
        self._msg_queue: queue.Queue = queue.Queue()
        self._running: bool = False
        self._run_history: list = []       # 실행 시각 문자열 목록
        self._log_lines: list = []
        self._run_records: list = []       # {success, fail, datetime}
        self._current_step: int = -1
        self._schedules: list = []         # [(BooleanVar, StringVar), ...]
        self._last_triggered_time: str = ""

        cfg = load_config()
        self._accounts: list = list(cfg.get("accounts", []))
        self._account_idx: int = int(cfg.get("account_index", 0))
        if self._account_idx >= len(self._accounts):
            self._account_idx = 0

        self._apply_dark_theme()
        self._build_ui(cfg)
        self._load_log_data()
        self._start_scheduler()
        self._poll_queue()
        if self._win is not None:
            self._win.protocol("WM_DELETE_WINDOW", self._on_close)

    # ------------------------------------------------------------------
    # 다크 테마
    # ------------------------------------------------------------------

    def _apply_dark_theme(self) -> None:
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure(".",
            background=C["bg"], foreground=C["fg"],
            fieldbackground=C["input"], bordercolor=C["bg"],
            darkcolor=NEU_DARK, lightcolor=NEU_LIGHT,
            troughcolor=C["input"],
            selectbackground=C["sel"], selectforeground=C["fg_bright"],
            insertcolor=C["fg"], relief="flat",
        )
        s.configure("TFrame", background=C["bg"])
        s.configure("Panel.TFrame", background=C["panel"])
        s.configure("Card.TFrame",
            background=C["bg"], relief="raised", borderwidth=5,
            lightcolor=NEU_LIGHT, darkcolor=NEU_DARK,
        )
        s.configure("TLabel",
            background=C["bg"], foreground=C["fg"], font=("Segoe UI", 9))
        s.configure("Panel.TLabel",
            background=C["panel"], foreground=C["fg"], font=("Segoe UI", 9))
        s.configure("Dim.TLabel",
            background=C["panel"], foreground=C["fg_dim"], font=("Segoe UI", 8))
        s.configure("TLabelframe",
            background=C["bg"], bordercolor=NEU_DARK,
            darkcolor=NEU_DARK, lightcolor=NEU_LIGHT, relief="raised", borderwidth=4,
        )
        s.configure("TLabelframe.Label",
            background=C["bg"], foreground=C["accent"],
            font=("Segoe UI", 9, "bold"),
        )
        s.configure("TEntry",
            fieldbackground=C["input"], foreground=C["fg"],
            bordercolor=NEU_DARK, lightcolor=NEU_DARK, darkcolor=NEU_LIGHT,
            insertcolor=C["fg"], relief="sunken", borderwidth=3, padding=(4, 3),
        )
        s.map("TEntry",
            fieldbackground=[("readonly", C["panel"])],
            lightcolor=[("focus", C["accent"])],
            darkcolor=[("focus", NEU_LIGHT)],
        )
        s.configure("TCheckbutton",
            background=C["bg"], foreground=C["fg"],
            focuscolor=C["bg"], font=("Segoe UI", 9),
        )
        s.map("TCheckbutton",
            background=[("active", C["bg"])],
            foreground=[("active", C["fg_bright"])],
        )
        s.configure("TButton",
            background=C["bg"], foreground=C["fg"],
            bordercolor=C["bg"], darkcolor=NEU_DARK, lightcolor=NEU_LIGHT,
            relief="raised", borderwidth=4, padding=(10, 5), font=("Segoe UI", 9),
        )
        s.map("TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", NEU_LIGHT), ("active", NEU_DARK)],
            lightcolor=[("pressed", NEU_DARK), ("active", NEU_LIGHT)],
            foreground=[("active", C["fg_bright"])],
        )
        s.configure("Start.TButton",
            background=C["start"], foreground="#ffffff",
            darkcolor="#1d5238", lightcolor="#52b88a",
            relief="raised", borderwidth=4, font=("Segoe UI", 10, "bold"),
        )
        s.map("Start.TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", "#52b88a")],
            lightcolor=[("pressed", "#1d5238")],
            background=[("active", C["start_hl"]), ("pressed", C["start"])],
        )
        s.configure("Stop.TButton",
            background=C["stop"], foreground="#ffffff",
            darkcolor="#7a1f1f", lightcolor="#e87070",
            relief="raised", borderwidth=4, font=("Segoe UI", 9, "bold"),
        )
        s.map("Stop.TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", "#e87070")],
            lightcolor=[("pressed", "#7a1f1f")],
            background=[("active", C["stop_hl"]), ("pressed", C["stop"])],
        )
        s.configure("TScrollbar",
            background=C["bg"], troughcolor=C["input"],
            bordercolor=C["bg"], darkcolor=NEU_DARK, lightcolor=NEU_LIGHT,
            arrowcolor=C["fg_dim"], relief="raised", borderwidth=2,
        )
        s.map("TScrollbar", background=[("active", C["panel2"])])
        s.configure("TSeparator", background=NEU_DARK)
        s.configure("Treeview",
            background=C["input"], foreground=C["fg"],
            fieldbackground=C["input"], bordercolor=NEU_DARK,
            rowheight=26, font=("Segoe UI", 9),
        )
        s.configure("Treeview.Heading",
            background=C["bg"], foreground=C["fg_dim"],
            bordercolor=NEU_DARK, darkcolor=NEU_DARK, lightcolor=NEU_LIGHT,
            font=("Segoe UI", 9, "bold"), relief="raised", borderwidth=2,
        )
        s.map("Treeview",
            background=[("selected", C["sel"])],
            foreground=[("selected", C["fg_bright"])],
        )
        s.map("Treeview.Heading", background=[("active", C["panel2"])])

    # ------------------------------------------------------------------
    # 카드 헬퍼
    # ------------------------------------------------------------------

    def _make_card(self, parent: tk.Widget, title: str = "") -> tk.Frame:
        outer = ttk.Frame(parent, style="Card.TFrame", padding=3)
        pad = tk.Frame(outer, bg=C["bg"])
        pad.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
        if title:
            tk.Label(pad, text=title.upper(),
                     bg=C["bg"], fg=C["fg_dim"],
                     font=("Segoe UI", 7, "bold")).pack(anchor=tk.W, pady=(0, 8))
            body = tk.Frame(pad, bg=C["bg"])
            body.pack(fill=tk.BOTH, expand=True)
        else:
            body = pad
        body._outer = outer
        return body

    # ------------------------------------------------------------------
    # UI 구성
    # ------------------------------------------------------------------

    def _build_ui(self, cfg: dict) -> None:
        G = 8

        # ── 상태바 ────────────────────────────────────────────────
        self._status_var = tk.StringVar(value="대기 중")
        tk.Label(self, textvariable=self._status_var,
                 bg=C["panel2"], fg=C["fg_dim"],
                 font=("Segoe UI", 8), anchor=tk.W,
                 padx=10, pady=3).pack(fill=tk.X, side=tk.BOTTOM)

        # ── 메인 컨테이너 ─────────────────────────────────────────
        main = tk.Frame(self, bg=C["bg"])
        main.pack(fill=tk.BOTH, expand=True, padx=G, pady=G)

        # ── 상단 행: 실행 제어(좌) + 사이트·계정(우) ──────────────
        top_row = tk.Frame(main, bg=C["bg"])
        top_row.pack(fill=tk.X, pady=(0, G))

        # ─── 실행 제어 카드 (좌) ──────────────────────────────────
        ca = self._make_card(top_row)
        ca._outer.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, G))

        tk.Label(ca, text="실행 제어",
                 bg=C["panel"], fg=C["accent"],
                 font=("Segoe UI", 8, "bold")).pack(anchor=tk.W, pady=(0, 12))

        btn_row = tk.Frame(ca, bg=C["panel"])
        btn_row.pack(fill=tk.X)
        self._run_btn = ttk.Button(btn_row, text="▶  실행",
                                   style="Start.TButton",
                                   command=self._run_submit)
        self._run_btn.pack(side=tk.LEFT, padx=(0, 6))
        self._stop_btn = ttk.Button(btn_row, text="■  중지",
                                    style="Stop.TButton",
                                    command=self._stop_submit,
                                    state=tk.DISABLED)
        self._stop_btn.pack(side=tk.LEFT)

        tk.Frame(ca, bg=C["border"], height=1).pack(fill=tk.X, pady=(12, 6))

        tk.Label(ca, text="예약 실행 시각  (HH:MM)",
                 font=("Segoe UI", 8), fg=C["fg_dim"],
                 bg=C["panel"]).pack(anchor=tk.W, pady=(0, 3))

        sched_times   = cfg.get("schedule_times",   ["", "", "", ""])
        sched_enabled = cfg.get("schedule_enabled", [False, False, False, False])
        while len(sched_times)   < 4: sched_times.append("")
        while len(sched_enabled) < 4: sched_enabled.append(False)

        for row_idx in range(2):
            srow = tk.Frame(ca, bg=C["panel"])
            srow.pack(fill=tk.X, pady=1)
            for col_idx in range(2):
                i = row_idx * 2 + col_idx
                ev = tk.BooleanVar(value=bool(sched_enabled[i]))
                tv = tk.StringVar(value=str(sched_times[i]))
                ttk.Checkbutton(srow, variable=ev,
                                style="TCheckbutton").pack(side=tk.LEFT)
                tk.Entry(srow,
                         textvariable=tv,
                         font=("Consolas", 10),
                         fg=C["yellow"], bg=C["input"],
                         insertbackground=C["yellow"],
                         relief=tk.FLAT, justify=tk.CENTER,
                         width=6).pack(side=tk.LEFT, padx=(2, 14))
                self._schedules.append((ev, tv))

        tk.Frame(ca, bg=C["border"], height=1).pack(fill=tk.X, pady=(8, 6))

        tk.Label(ca, text="최근 실행:",
                 font=("Segoe UI", 8), fg=C["fg_dim"],
                 bg=C["panel"]).pack(anchor=tk.W)
        self._history_var = tk.StringVar(value="없음")
        tk.Label(ca, textvariable=self._history_var,
                 font=("Consolas", 8), fg=C["accent"], bg=C["panel"],
                 wraplength=210, justify=tk.LEFT).pack(anchor=tk.W, padx=2, pady=(2, 0))

        # ─── 우측 컬럼 (사이트 설정 + 계정 설정) ─────────────────
        right_col = tk.Frame(top_row, bg=C["bg"])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ─── 사이트 설정 카드 ──────────────────────────────────────
        cb = self._make_card(right_col, "사이트 설정")
        cb._outer.pack(fill=tk.X, pady=(0, G))
        cb.columnconfigure(1, weight=1)
        cb.columnconfigure(3, weight=1)

        if not SELENIUM_OK:
            ttk.Label(cb, text="⚠  selenium 미설치 — 자동 클릭 불가",
                      foreground=C["error"],
                      style="Panel.TLabel").grid(
                row=0, column=0, columnspan=4, sticky=tk.W, pady=(0, 6))

        saved_urls = cfg.get("site_urls", DEFAULT_CONFIG["site_urls"])
        if isinstance(saved_urls, str):
            saved_urls = [saved_urls] + [""] * 9
        while len(saved_urls) < 10:
            saved_urls.append("")
        if self._shared_site_url_vars is not None:
            self._site_url_vars = self._shared_site_url_vars
        else:
            self._site_url_vars = [tk.StringVar(value=saved_urls[i]) for i in range(10)]
        self._site_url_entries = []

        for pr, (li, ri) in enumerate([(0,1),(2,3),(4,5),(6,7),(8,9)], start=1):
            ttk.Label(cb, text=f"사이트 {li+1}", style="Panel.TLabel").grid(
                row=pr, column=0, sticky=tk.W, pady=2)
            e_l = tk.Entry(cb, textvariable=self._site_url_vars[li],
                           bg=C["input"], fg=C["fg"], insertbackground=C["fg"],
                           relief=tk.FLAT, bd=1)
            e_l.grid(row=pr, column=1, sticky=tk.EW, padx=(6, 14), pady=2)
            self._site_url_entries.append(e_l)
            ttk.Label(cb, text=f"사이트 {ri+1}", style="Panel.TLabel").grid(
                row=pr, column=2, sticky=tk.W, pady=2)
            e_r = tk.Entry(cb, textvariable=self._site_url_vars[ri],
                           bg=C["input"], fg=C["fg"], insertbackground=C["fg"],
                           relief=tk.FLAT, bd=1)
            e_r.grid(row=pr, column=3, sticky=tk.EW, padx=(6, 0), pady=2)
            self._site_url_entries.append(e_r)

        ttk.Label(cb, text="Chrome 포트", style="Panel.TLabel").grid(
            row=6, column=0, sticky=tk.W, pady=(6, 2))
        pf = ttk.Frame(cb, style="Panel.TFrame")
        pf.grid(row=6, column=1, sticky=tk.W, padx=(6, 14), pady=(6, 2))
        self._port_var = tk.StringVar(value=str(cfg.get("chrome_port", 9222)))
        ttk.Entry(pf, textvariable=self._port_var, width=7).pack(side=tk.LEFT)
        ttk.Label(pf, text="  기본 9222", style="Dim.TLabel").pack(side=tk.LEFT)
        ttk.Button(cb, text="접속 테스트", command=self._test_connections).grid(
            row=6, column=2, columnspan=2, sticky=tk.EW, padx=(6, 0), pady=(6, 2))

        # ─── 계정 설정 카드 ────────────────────────────────────────
        cc = self._make_card(right_col, "계정 설정")
        cc._outer.pack(fill=tk.X)
        cc.columnconfigure(1, weight=1)

        ttk.Label(cc, text="현재 계정", style="Panel.TLabel").grid(
            row=0, column=0, sticky=tk.W, pady=3)
        self._current_acct_var = tk.StringVar()
        ttk.Label(cc, textvariable=self._current_acct_var,
                  style="Panel.TLabel",
                  foreground=C["accent"]).grid(
            row=0, column=1, sticky=tk.W, padx=(8, 0), pady=3)
        ttk.Button(cc, text="계정 관리",
                   command=self._open_account_manager).grid(
            row=0, column=2, sticky=tk.E, pady=3)
        self._update_current_acct_label()

        tk.Frame(cc, bg=C["border"], height=1).grid(
            row=1, column=0, columnspan=3, sticky=tk.EW, pady=(6, 6))

        ctrl_row = tk.Frame(cc, bg=C["panel"])
        ctrl_row.grid(row=2, column=0, columnspan=3, sticky=tk.W)
        ttk.Button(ctrl_row, text="로그 지우기",
                   command=self._clear_log).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(ctrl_row, text="텍스트 저장",
                   command=self._export_text).pack(side=tk.LEFT)

        # ─── 진행 단계 카드 ────────────────────────────────────────
        cstep = self._make_card(main, "진행 단계")
        cstep._outer.pack(fill=tk.X, pady=(0, G))

        self._step_circles: list = []
        self._step_labels_w: list = []

        step_row = tk.Frame(cstep, bg=C["panel"])
        step_row.pack(fill=tk.X)

        for i, (name, _) in enumerate(STEPS):
            circ = tk.Label(step_row, text="○",
                            bg=C["panel"], fg=C["border"],
                            font=("Segoe UI", 14))
            circ.pack(side=tk.LEFT)
            lbl = tk.Label(step_row, text=name,
                           bg=C["panel"], fg=C["fg_dim"],
                           font=("Segoe UI", 8))
            lbl.pack(side=tk.LEFT, padx=(1, 0))
            self._step_circles.append(circ)
            self._step_labels_w.append(lbl)
            if i < len(STEPS) - 1:
                tk.Label(step_row, text=" → ",
                         bg=C["panel"], fg=C["border"],
                         font=("Segoe UI", 9)).pack(side=tk.LEFT)

        # ─── 실행 로그 카드 ────────────────────────────────────────
        cd = self._make_card(main, "실행 로그")
        cd._outer.pack(fill=tk.BOTH, expand=True)

        self._log = tk.Text(
            cd, state=tk.DISABLED,
            background=C["log_bg"], foreground=C["fg"],
            insertbackground=C["fg"],
            relief=tk.FLAT, wrap=tk.WORD,
            font=("Consolas", 10),
            selectbackground=C["sel"],
            padx=6, pady=4,
        )
        sb = ttk.Scrollbar(cd, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._log.pack(fill=tk.BOTH, expand=True)

        self._log.tag_configure("system", foreground=C["system"],
                                font=("Consolas", 10, "italic"))
        self._log.tag_configure("error",  foreground=C["error"])
        self._log.tag_configure("ok",     foreground=C["ok"],
                                font=("Consolas", 10, "italic"))

    # ------------------------------------------------------------------
    # 단계 진행 표시
    # ------------------------------------------------------------------

    def _detect_step(self, text: str) -> int:
        """텍스트에서 해당 단계 인덱스 반환. 미감지 시 -1."""
        for i, (_, keywords) in enumerate(STEPS):
            for kw in keywords:
                if kw.lower() in text.lower():
                    return i
        return -1

    def _update_step_display(self, step: int, done: bool = False, error: bool = False) -> None:
        self._current_step = step
        for i, (circ, lbl) in enumerate(
                zip(self._step_circles, self._step_labels_w)):
            if error and i == step:
                circ.config(text="✕", fg=C["error"])
                lbl.config(fg=C["error"])
            elif done or i < step:
                circ.config(text="●", fg=C["ok"])
                lbl.config(fg=C["ok"])
            elif i == step:
                circ.config(text="●", fg=C["accent"])
                lbl.config(fg=C["fg_bright"])
            else:
                circ.config(text="○", fg=C["border"])
                lbl.config(fg=C["fg_dim"])

    def _reset_steps(self) -> None:
        self._current_step = -1
        for circ, lbl in zip(self._step_circles, self._step_labels_w):
            circ.config(text="○", fg=C["border"])
            lbl.config(fg=C["fg_dim"])

    # ------------------------------------------------------------------
    # 예약 실행
    # ------------------------------------------------------------------

    def _start_scheduler(self) -> None:
        self._last_triggered_time = ""
        self._check_schedule()

    def _check_schedule(self) -> None:
        if not self._running:
            now = datetime.datetime.now().strftime("%H:%M")
            if now != self._last_triggered_time:
                for ev, tv in self._schedules:
                    t = tv.get().strip()
                    if ev.get() and t == now:
                        self._last_triggered_time = now
                        self._append_log(f"⏰ 예약 실행: {now}", tag="system")
                        self._run_submit()
                        break
        self.after(10000, self._check_schedule)

    # ------------------------------------------------------------------
    # 실행 / 중지
    # ------------------------------------------------------------------

    def _run_submit(self) -> None:
        if self._running:
            return
        if not SELENIUM_OK:
            self._append_log("⚠ selenium 미설치 — 자동 클릭 불가", tag="error")
            return
        try:
            port = int(self._port_var.get())
        except ValueError:
            port = 9222
        site_urls = [v.get().strip() for v in self._site_url_vars if v.get().strip()]
        if not site_urls:
            self._append_log("⚠ 사이트 주소를 1개 이상 입력하세요.", tag="error")
            return

        self._auto_cancel_event.clear()
        self._running = True
        self._run_btn.config(state=tk.DISABLED)
        self._stop_btn.config(state=tk.NORMAL)
        self._status_var.set("실행 중...")
        self._reset_steps()

        now = datetime.datetime.now().strftime("%H:%M:%S")
        self._run_history.insert(0, now)
        self._run_history = self._run_history[:8]
        self._history_var.set("\n".join(self._run_history))

        threading.Thread(
            target=self._run_auto_input,
            args=(port, site_urls),
            daemon=True,
        ).start()

    def _stop_submit(self) -> None:
        self._auto_cancel_event.set()
        self._status_var.set("중지 요청...")

    def _run_auto_input(self, port: int, site_urls: list) -> None:
        accounts = [a for a in self._accounts if a.get("enabled", True)]
        if not accounts:
            self._msg_queue.put(("error",
                "⚠ 활성화된 계정이 없습니다. 계정 관리에서 ☑ 체크하세요."))
            self._msg_queue.put(("__done__", False))
            return

        total = len(accounts)
        self._msg_queue.put(("system",
            f"━━ 실행 시작 | 활성 계정 {total}개 ━━"))

        success_count = 0
        fail_count = 0

        for i, acct in enumerate(accounts):
            if self._auto_cancel_event.is_set():
                self._msg_queue.put(("system", "── 취소됨 ──"))
                break

            acct_label = f"[{i+1}/{total}] {acct['email']}"
            self._msg_queue.put(("system", f"→ {acct_label} 처리 중..."))

            def _status(msg: str) -> None:
                self._msg_queue.put(("system", f"  · {msg}"))

            try:
                click_no_more(
                    port, site_urls,
                    email=acct["email"],
                    password=acct["password"],
                    status_cb=_status,
                    cancel_event=self._auto_cancel_event,
                )
                self._msg_queue.put(("ok", f"✓ 완료: {acct_label}"))
                success_count += 1
            except AutoCancelled:
                self._msg_queue.put(("system", "── 취소됨 ──"))
                break
            except Exception as e:
                self._msg_queue.put(("error", f"✗ 실패: {acct_label} [{type(e).__name__}] — {e}"))
                fail_count += 1

        tag = "ok" if fail_count == 0 else "system"
        self._msg_queue.put((
            tag,
            f"━━ 완료 | 성공 {success_count} / 실패 {fail_count} / 총 {total} ━━",
        ))
        self._run_records.append({
            "success":  success_count,
            "fail":     fail_count,
            "datetime": datetime.datetime.now().isoformat(timespec="seconds"),
        })
        self._msg_queue.put(("__done__", fail_count == 0))

    # ------------------------------------------------------------------
    # 큐 폴링
    # ------------------------------------------------------------------

    def _poll_queue(self) -> None:
        try:
            while True:
                tag, data = self._msg_queue.get_nowait()
                if tag == "__done__":
                    self._on_submit_done(data)
                else:
                    self._append_log(data, tag=tag)
        except queue.Empty:
            pass
        self.after(200, self._poll_queue)

    def _on_submit_done(self, success: bool = True) -> None:
        self._running = False
        self._run_btn.config(state=tk.NORMAL)
        self._stop_btn.config(state=tk.DISABLED)
        self._update_step_display(len(STEPS) - 1, done=True)
        self._status_var.set(
            f"완료  |  {datetime.datetime.now().strftime('%H:%M:%S')}")

    # ------------------------------------------------------------------
    # 로그
    # ------------------------------------------------------------------

    def _append_log(self, text: str, tag: str = "") -> None:
        if tag == "system" and self._running:
            step = self._detect_step(text)
            if step >= 0:
                self._update_step_display(step)
        self._log_lines.append({"text": text, "tag": tag})
        if len(self._log_lines) > 2000:
            self._log_lines = self._log_lines[-2000:]
        self._log.config(state=tk.NORMAL)
        self._log.insert(tk.END, text + "\n", tag if tag else ())
        self._log.see(tk.END)
        self._log.config(state=tk.DISABLED)

    def _clear_log(self) -> None:
        self._log_lines.clear()
        self._log.config(state=tk.NORMAL)
        self._log.delete("1.0", tk.END)
        self._log.config(state=tk.DISABLED)

    # ------------------------------------------------------------------
    # 로그 영구 저장 / 복원 / 내보내기
    # ------------------------------------------------------------------

    def _load_log_data(self) -> None:
        if not LOG_DATA_PATH.exists():
            return
        try:
            with open(LOG_DATA_PATH, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return
        lines = data.get("log_lines", [])
        self._log_lines = lines[-2000:]
        if self._log_lines:
            self._log.config(state=tk.NORMAL)
            for item in self._log_lines:
                tag = item.get("tag", "")
                self._log.insert(tk.END, item.get("text", "") + "\n",
                                 tag if tag else ())
            self._log.see(tk.END)
            self._log.config(state=tk.DISABLED)
        self._run_records = data.get("run_records", [])
        history = data.get("run_history", [])
        if history:
            self._run_history = history[:8]
            self._history_var.set("\n".join(self._run_history))

    def _save_log_data(self) -> None:
        try:
            data = {
                "run_records":  self._run_records,
                "log_lines":    self._log_lines[-2000:],
                "run_history":  self._run_history,
            }
            with open(LOG_DATA_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.warning(f"로그 데이터 저장 실패: {e}")

    def _export_text(self) -> None:
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")],
            initialfile=f"보너스픽_{now_str}.txt",
            title="텍스트 파일로 저장",
        )
        if not path:
            return
        sep = "═" * 60
        out = [sep, f"  실행 기록  (총 {len(self._run_records)}건)", sep]
        for i, r in enumerate(self._run_records, 1):
            dt = r.get("datetime", "")[:19].replace("T", " ")
            out.append(
                f"  {i:>3}.  성공 {r.get('success', 0)} / 실패 {r.get('fail', 0)}"
                f"   {dt}")
        out += ["", sep, "  실행 로그", sep]
        for item in self._log_lines:
            out.append(item.get("text", ""))
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(out))
            self._append_log(f"✓ 저장 완료: {path}", tag="ok")
        except Exception as e:
            messagebox.showerror("저장 실패", str(e), parent=self)

    # ------------------------------------------------------------------
    # 계정 관리
    # ------------------------------------------------------------------

    def _update_current_acct_label(self) -> None:
        if not self._accounts:
            self._current_acct_var.set("등록된 계정 없음 — [계정 관리]에서 추가하세요")
        else:
            enabled = [a for a in self._accounts if a.get("enabled", True)]
            self._current_acct_var.set(
                f"총 {len(self._accounts)}개  |  활성 {len(enabled)}개")

    def _test_connections(self) -> None:
        """입력된 사이트 URL 접속 가능 여부를 테스트. 실패 URL은 빨간색으로 표시."""
        import urllib.request
        pairs = [(self._site_url_entries[i], v.get().strip())
                 for i, v in enumerate(self._site_url_vars) if v.get().strip()]
        if not pairs:
            messagebox.showinfo("접속 테스트", "입력된 사이트 URL이 없습니다.")
            return

        # 테스트 전 모든 대상 배경 초기화
        for entry, _ in pairs:
            entry.config(bg=C["input"])

        def _run():
            lines = []
            for entry, url in pairs:
                try:
                    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                    with urllib.request.urlopen(req, timeout=8):
                        lines.append(f"✓  {url}")
                        self.after(0, lambda e=entry: e.config(bg=C["input"]))
                except Exception as ex:
                    lines.append(f"✗  {url}\n     {ex}")
                    self.after(0, lambda e=entry: e.config(bg="#5c1a1a"))
            self.after(0, lambda m="\n\n".join(lines): messagebox.showinfo("접속 테스트 결과", m))

        messagebox.showinfo("접속 테스트", f"{len(pairs)}개 사이트 접속 테스트 중...\n완료되면 결과 팝업이 표시됩니다.")
        threading.Thread(target=_run, daemon=True).start()

    def _open_account_manager(self) -> None:
        def on_save(accounts, current_idx):
            self._accounts = accounts
            self._account_idx = current_idx
            self._update_current_acct_label()
            save_config(self._get_current_config())
        AccountManagerDialog(self, self._accounts, self._account_idx, on_save)

    # ------------------------------------------------------------------
    # 설정 / 종료
    # ------------------------------------------------------------------

    def _get_current_config(self) -> dict:
        try:
            port = int(self._port_var.get())
        except ValueError:
            port = 9222
        return {
            "site_urls":        [v.get().strip() for v in self._site_url_vars],
            "chrome_port":      port,
            "accounts":         self._accounts,
            "account_index":    self._account_idx,
            "schedule_times":   [tv.get().strip() for _, tv in self._schedules],
            "schedule_enabled": [bool(ev.get()) for ev, _ in self._schedules],
        }

    def _on_close(self) -> None:
        self._auto_cancel_event.set()
        self._save_log_data()
        save_config(self._get_current_config())
        if self._win is not None:
            self._win.destroy()


# ---------------------------------------------------------------------------
# 계정 관리 다이얼로그
# ---------------------------------------------------------------------------

class _AccountEditDialog(tk.Toplevel):
    def __init__(self, parent, account: dict = None) -> None:
        super().__init__(parent)
        self.title("계정 추가" if account is None else "계정 수정")
        self.resizable(False, False)
        self.configure(bg=C["bg"])
        self.grab_set()
        self.result: Optional[dict] = None
        self.columnconfigure(1, weight=1)

        fields = [
            ("이메일",   "email",    False),
            ("비밀번호", "password", True),
            ("비고",     "memo",     False),
        ]
        self._vars = {}
        for r, (label, key, is_pw) in enumerate(fields):
            ttk.Label(self, text=f"{label}:", style="TLabel").grid(
                row=r, column=0, padx=14, pady=6, sticky=tk.W)
            var = tk.StringVar(value=account.get(key, "") if account else "")
            self._vars[key] = var
            entry = ttk.Entry(self, textvariable=var, width=34,
                              show="*" if is_pw else "")
            entry.grid(row=r, column=1, padx=(0, 14), pady=6, sticky=tk.EW)
            if is_pw:
                self._pw_entry = entry

        self._show_pw = tk.BooleanVar(value=False)
        ttk.Checkbutton(self, text="비밀번호 표시", variable=self._show_pw,
                        command=lambda: self._pw_entry.config(
                            show="" if self._show_pw.get() else "*")).grid(
            row=1, column=2, padx=(0, 14))

        self._enabled_var = tk.BooleanVar(
            value=account.get("enabled", True) if account else True)
        ttk.Checkbutton(self, text="이 계정 사용",
                        variable=self._enabled_var).grid(
            row=3, column=1, sticky=tk.W, padx=(0, 14), pady=(0, 6))

        btn_frame = ttk.Frame(self)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=12)
        ttk.Button(btn_frame, text="확인", style="Start.TButton",
                   command=self._ok).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frame, text="취소",
                   command=self.destroy).pack(side=tk.LEFT, padx=6)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _ok(self) -> None:
        email = self._vars["email"].get().strip()
        if not email:
            messagebox.showwarning("입력 오류", "이메일을 입력하세요.", parent=self)
            return
        self.result = {
            "email":    email,
            "password": self._vars["password"].get(),
            "memo":     self._vars["memo"].get().strip(),
            "enabled":  self._enabled_var.get(),
        }
        self.destroy()


class AccountManagerDialog(tk.Toplevel):
    def __init__(self, parent, accounts: list, current_idx: int, on_save) -> None:
        super().__init__(parent)
        self.title("계정 목록 관리")
        self.resizable(True, True)
        self.minsize(580, 460)
        self.configure(bg=C["bg"])
        self.grab_set()

        self._accounts: list = list(accounts)
        self._current_idx: int = current_idx
        self._on_save = on_save

        self._build_ui()
        self._refresh_list()

    def _build_ui(self) -> None:
        ttk.Label(self,
                  text="  ☑ 클릭 → 사용/미사용 전환    ✕ 클릭 → 삭제    더블클릭 → 수정",
                  foreground=C["fg_dim"],
                  font=("Segoe UI", 8)).pack(padx=10, pady=(8, 2), anchor=tk.W)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        cols = ("check", "no", "email", "password", "memo", "del")
        self._tree = ttk.Treeview(tree_frame, columns=cols,
                                  show="headings", height=16)
        self._tree.heading("check",    text="사용")
        self._tree.heading("no",       text="No.")
        self._tree.heading("email",    text="이메일")
        self._tree.heading("password", text="비밀번호")
        self._tree.heading("memo",     text="비고")
        self._tree.heading("del",      text="삭제")
        self._tree.column("check",    width=45,  anchor="center", stretch=False)
        self._tree.column("no",       width=45,  anchor="center", stretch=False)
        self._tree.column("email",    width=200)
        self._tree.column("password", width=80,  stretch=False)
        self._tree.column("memo",     width=130)
        self._tree.column("del",      width=45,  anchor="center", stretch=False)

        sb = ttk.Scrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._tree.bind("<Double-1>",       self._on_double_click)
        self._tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self._tree.tag_configure("current",  foreground=C["accent"])
        self._tree.tag_configure("disabled", foreground="#555555")

        btn_frame = ttk.Frame(self)
        btn_frame.pack(padx=10, pady=(4, 10))
        ttk.Button(btn_frame, text="추가",
                   command=self._add).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="수정",
                   command=self._edit).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="전체 선택",
                   command=self._enable_all).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="전체 해제",
                   command=self._disable_all).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="엑셀 불러오기",
                   command=self._import_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="엑셀 저장",
                   command=self._export_excel).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="저장 후 닫기", style="Start.TButton",
                   command=self._save_close).pack(side=tk.LEFT, padx=3)

    def _refresh_list(self) -> None:
        for item in self._tree.get_children():
            self._tree.delete(item)
        for i, acct in enumerate(self._accounts):
            enabled = acct.get("enabled", True)
            check  = "☑" if enabled else "☐"
            marker = "▶" if i == self._current_idx else str(i + 1)
            masked = "•" * min(len(acct.get("password", "")), 8) or "(없음)"
            tags = []
            if i == self._current_idx:
                tags.append("current")
            if not enabled:
                tags.append("disabled")
            self._tree.insert("", tk.END, iid=str(i),
                              values=(check, marker, acct.get("email", ""),
                                      masked, acct.get("memo", ""), "✕"),
                              tags=tuple(tags))
        if self._accounts and 0 <= self._current_idx < len(self._accounts):
            self._tree.selection_set(str(self._current_idx))
            self._tree.see(str(self._current_idx))

    def _on_tree_click(self, event) -> None:
        if self._tree.identify_region(event.x, event.y) != "cell":
            return
        col    = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        idx = int(row_id)
        if col == "#1":
            self._toggle_enabled(idx)
        elif col == "#6":
            self._delete_by_idx(idx)

    def _on_double_click(self, event) -> None:
        if self._tree.identify_column(event.x) in ("#1", "#6"):
            return
        self._edit()

    def _toggle_enabled(self, idx: int) -> None:
        self._accounts[idx]["enabled"] = not self._accounts[idx].get("enabled", True)
        self._refresh_list()

    def _delete_by_idx(self, idx: int) -> None:
        if not messagebox.askyesno(
                "삭제 확인",
                f"계정 {idx+1} ({self._accounts[idx]['email']})을 삭제하시겠습니까?",
                parent=self):
            return
        self._accounts.pop(idx)
        if self._current_idx >= len(self._accounts):
            self._current_idx = max(0, len(self._accounts) - 1)
        self._refresh_list()

    def _enable_all(self) -> None:
        for a in self._accounts:
            a["enabled"] = True
        self._refresh_list()

    def _disable_all(self) -> None:
        for a in self._accounts:
            a["enabled"] = False
        self._refresh_list()

    def _selected_idx(self) -> Optional[int]:
        sel = self._tree.selection()
        return int(sel[0]) if sel else None

    def _add(self) -> None:
        if len(self._accounts) >= 50:
            messagebox.showwarning("최대 50개", "계정은 최대 50개까지 등록 가능합니다.",
                                   parent=self)
            return
        dlg = _AccountEditDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self._accounts.append(dlg.result)
            self._refresh_list()

    def _edit(self) -> None:
        idx = self._selected_idx()
        if idx is None:
            return
        dlg = _AccountEditDialog(self, self._accounts[idx])
        self.wait_window(dlg)
        if dlg.result:
            self._accounts[idx] = dlg.result
            self._refresh_list()

    def _import_excel(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")],
            title="계정 목록 엑셀 파일 선택",
            parent=self,
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            messagebox.showerror("불러오기 실패", str(e), parent=self)
            return

        if not rows:
            messagebox.showwarning("내용 없음", "엑셀 파일이 비어 있습니다.", parent=self)
            return

        # 헤더 행 스킵 (첫 셀이 '이메일' 또는 'email'이면 헤더로 판단)
        start = 0
        first = str(rows[0][0] or "").strip().lower()
        if first in ("이메일", "email", "e-mail"):
            start = 1

        new_accounts = []
        skipped = 0
        for row in rows[start:]:
            if not row or not row[0]:
                continue
            email = str(row[0]).strip()
            if not email:
                continue
            pw    = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            memo  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            enabled_raw = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else "true"
            enabled = enabled_raw not in ("false", "x", "✕", "0", "n", "no", "미사용")
            new_accounts.append({"email": email, "password": pw,
                                  "memo": memo, "enabled": enabled})

        if not new_accounts:
            messagebox.showwarning("계정 없음", "불러올 계정이 없습니다.", parent=self)
            return

        mode = messagebox.askyesnocancel(
            "불러오기 방식",
            f"엑셀에서 계정 {len(new_accounts)}개를 찾았습니다.\n\n"
            "[예] 기존 목록에 추가\n"
            "[아니오] 기존 목록을 대체\n"
            "[취소] 취소",
            parent=self,
        )
        if mode is None:
            return
        if mode:
            self._accounts.extend(new_accounts)
        else:
            self._accounts = new_accounts
            self._current_idx = 0
        self._refresh_list()
        messagebox.showinfo("완료", f"계정 {len(new_accounts)}개를 불러왔습니다.", parent=self)

    def _export_excel(self) -> None:
        if not self._accounts:
            messagebox.showwarning("계정 없음", "내보낼 계정이 없습니다.", parent=self)
            return
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            initialfile=f"보너스픽_계정_{now_str}.xlsx",
            title="계정 목록 저장",
            parent=self,
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "계정목록"

            headers = ["이메일", "비밀번호", "비고", "사용여부"]
            header_fill = PatternFill("solid", fgColor="2D6A2D")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for r, acct in enumerate(self._accounts, 2):
                ws.cell(row=r, column=1, value=acct.get("email", ""))
                ws.cell(row=r, column=2, value=acct.get("password", ""))
                ws.cell(row=r, column=3, value=acct.get("memo", ""))
                ws.cell(row=r, column=4, value="사용" if acct.get("enabled", True) else "미사용")

            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            wb.save(path)
        except Exception as e:
            messagebox.showerror("저장 실패", str(e), parent=self)
            return
        messagebox.showinfo("완료", f"계정 {len(self._accounts)}개를 저장했습니다.\n{path}", parent=self)

    def _save_close(self) -> None:
        self._on_save(self._accounts, self._current_idx)
        self.destroy()


# ---------------------------------------------------------------------------
# 진입점
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = BonusPickApp()
    app._win.mainloop()
