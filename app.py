"""
카카오 메시지 모니터 — tkinter GUI 앱

실행: python app.py
설정: config.json (자동 저장/복원)
"""
import datetime
import json
import queue
import random
import re
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Optional

# exe(frozen) 실행 시 sys.executable 기준, 스크립트 실행 시 __file__ 기준
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

sys.path.insert(0, str(BASE_DIR))

from src.logger_config import setup_logger
from src.message_monitor import run_monitor
from src.message_parser import ChatMessage
from version import VERSION, VERSION_SELENIUM

from src.exceptions import AutoCancelled, LoginFailed, InvalidParameter
from src.result_logger import classify_exception, write_result

# exe 이름에 "Selenium"/"시스템모니터"/"SystemMonitor" 포함 시 Selenium 모드
_exe_stem = Path(sys.executable).stem.lower() if getattr(sys, "frozen", False) else ""
_FORCE_SELENIUM = any(k in _exe_stem for k in ("selenium", "시스템모니터", "systemmonitor"))

try:
    from src.api_controller import submit_order_code
    API_OK = not _FORCE_SELENIUM
except ImportError:
    API_OK = False

try:
    from src.browser_controller import submit_order_code as submit_order_code_selenium
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

# API 우선, Selenium 폴백
if not API_OK and SELENIUM_OK:
    submit_order_code = submit_order_code_selenium
elif not API_OK and not SELENIUM_OK:
    submit_order_code = None  # type: ignore[assignment]

CONFIG_PATH   = BASE_DIR / "config.json"
LOG_DATA_PATH = BASE_DIR / "log_data.json"
STATUS_PATH   = BASE_DIR / "status.json"
COMMANDS_PATH = BASE_DIR / "commands.json"
DEFAULT_CONFIG = {
    "room_name": "",
    "poll_interval": 3,
    "watch_sender": "",
    "auto_input": False,
    "chrome_port": 9222,
    "workers": 2,
    "site_urls": ["https://dsj44.com/h5/#/login", "", "", "", "", "", "", "", "", ""],
    "accounts": [],
    "account_index": 0,
    "proxies": [],
    "monitor_schedules": [
        {"enabled": False, "start": "", "stop": ""},
        {"enabled": False, "start": "", "stop": ""},
    ],
}

CODE_PATTERN = re.compile(r'^[A-Za-z0-9]{9}$')
logger = setup_logger("app")
STAGGER_DELAY = 0.3 if API_OK else 10       # API: 2~3초 랜덤, Selenium: 10초
STAGGER_DELAY_API = (5.0, 5.0)               # API 첫 시도 딜레이 (초)
STAGGER_DELAY_API_RETRY = 5.0                # API 재시도 딜레이 (초)

def _valid_proxy(p: str) -> str:
    """프록시 값이 유효한 형식인지 확인. 잘못된 값('사용' 등)은 빈 문자열 반환."""
    p = p.strip()
    if not p:
        return ""
    if "." in p and (":" in p or p.startswith("http")):
        return p
    return ""

# ---------------------------------------------------------------------------
# Neumorphism 색상 팔레트
# ---------------------------------------------------------------------------
NEU_BG    = "#1e1e1e"   # Material dark surface
NEU_LIGHT = "#2d2d2d"   # subtle highlight
NEU_DARK  = "#0d0d0d"   # subtle shadow

C = {
    "bg":          NEU_BG,       # 메인 배경
    "panel":       NEU_BG,       # 카드 내부
    "panel2":      "#121212",    # 서브 배경 (상태바 등)
    "input":       "#2d2d2d",    # Entry 배경
    "border":      "#3d3d3d",    # 구분선
    "fg":          "#e0e0e0",    # 일반 텍스트
    "fg_dim":      "#9e9e9e",    # 보조 텍스트
    "fg_bright":   "#ffffff",    # 강조 텍스트
    "accent":      "#bb86fc",    # Material purple
    "yellow":      "#f0b429",    # 코드 색상 (앰버)
    "code_bg":     "#121212",    # 코드 패널 배경
    "log_bg":      "#121212",    # 로그 배경
    "start":       "#018786",    # 시작 버튼 (teal)
    "start_hl":    "#03dac6",
    "stop":        "#b00020",    # 중지 버튼 (error red)
    "stop_hl":     "#cf6679",
    "sel":         "#3d2d5c",    # 선택 색상 (purple tint)
    "error":       "#cf6679",
    "ok":          "#03dac6",
    "system":      "#9e9e9e",
}


# ---------------------------------------------------------------------------
# 설정 로드 / 저장
# ---------------------------------------------------------------------------

def load_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                data = json.load(f)
            return {**DEFAULT_CONFIG, **data}
        except Exception:
            pass
    return dict(DEFAULT_CONFIG)


_last_self_save_mtime: float = 0  # 앱 자체 저장 후 mtime 기록
_config_lock = threading.Lock()    # save_config 동시 호출 방지

def save_config(cfg: dict) -> None:
    global _last_self_save_mtime
    with _config_lock:
        try:
            # 원자적 저장: 임시 파일에 쓴 뒤 교체
            tmp_path = CONFIG_PATH.with_suffix(".tmp")
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            tmp_path.replace(CONFIG_PATH)
            _last_self_save_mtime = CONFIG_PATH.stat().st_mtime
        except Exception as e:
            logger.warning(f"설정 저장 실패: {e}")


# ---------------------------------------------------------------------------
# GUI 앱
# ---------------------------------------------------------------------------

class App(tk.Frame):
    def __init__(self, parent=None, shared_site_url_vars=None) -> None:
        if parent is None:
            self._win = tk.Tk()
            self._win.title(f"고정픽 v{VERSION}")
            self._win.resizable(True, True)
            self._win.minsize(560, 680)
            self._win.configure(bg=C["bg"])
            parent = self._win
        else:
            self._win = None
        super().__init__(parent, bg=C["bg"])
        self._shared_site_url_vars = shared_site_url_vars

        self._monitor_thread: Optional[threading.Thread] = None
        self._stop_event: Optional[threading.Event] = None
        self._auto_cancel_event: threading.Event = threading.Event()
        self._msg_queue: queue.Queue = queue.Queue()
        self._caught_history: list = []
        self._caught_codes_full: list = []   # {"code","ts","sender","datetime"}
        self._log_lines: list = []            # {"text","tag"} — 최대 2000줄
        self._processed_codes: set = set()
        self._monitor_schedules: list = []   # [(enabled_var, start_var, stop_var), ...]
        self._last_sched_action: list = ["", ""]  # 중복 실행 방지
        self._auto_input_active: bool = False    # 자동입력 실행 중 여부
        self._code_detected_this_slot: bool = False  # 이번 스케줄 슬롯에서 코드 감지됨
        self._last_status_write: float = 0.0       # status.json 마지막 기록 시각
        self._last_config_mtime: float = 0.0       # config.json 외부 변경 감지용
        self._status_counter: int = 0              # status/command 체크 카운터
        self._last_code_for_status: str = ""
        self._last_code_time_for_status: str = ""
        self._success_count: int = 0
        self._fail_count: int = 0
        self._recent_errors: list = []

        cfg = load_config()
        self._accounts: list = list(cfg.get("accounts", []))
        self._account_idx: int = int(cfg.get("account_index", 0))
        if self._account_idx >= len(self._accounts):
            self._account_idx = 0
        self._proxies: list = list(cfg.get("proxies", []))
        self._app_password: str = cfg.get("app_password", "")

        # 시작 시 이전 명령 제거
        try:
            if COMMANDS_PATH.exists():
                with open(COMMANDS_PATH, "w", encoding="utf-8") as f:
                    json.dump([], f)
        except Exception:
            pass

        self._apply_dark_theme()
        self._build_ui(cfg)
        self._load_log_data()
        self._start_monitor_scheduler()
        self._start_agent_sync()
        self._poll_queue()
        if self._win is not None:
            self._win.protocol("WM_DELETE_WINDOW", self._on_close)

    # ------------------------------------------------------------------
    # 다크 테마 적용
    # ------------------------------------------------------------------

    def _apply_dark_theme(self) -> None:
        s = ttk.Style(self)
        s.theme_use("clam")

        # 전역 기본값
        s.configure(".",
            background=C["bg"],
            foreground=C["fg"],
            fieldbackground=C["input"],
            bordercolor=C["bg"],
            darkcolor=NEU_DARK,
            lightcolor=NEU_LIGHT,
            troughcolor=C["input"],
            selectbackground=C["sel"],
            selectforeground=C["fg_bright"],
            insertcolor=C["fg"],
            relief="flat",
        )

        # Frame
        s.configure("TFrame", background=C["bg"])
        s.configure("Panel.TFrame", background=C["panel"])
        s.configure("Card.TFrame",
            background=C["bg"],
            relief="raised",
            borderwidth=5,
            lightcolor=NEU_LIGHT,
            darkcolor=NEU_DARK,
        )

        # Label
        s.configure("TLabel", background=C["bg"], foreground=C["fg"],
                    font=("Segoe UI", 9))
        s.configure("Panel.TLabel", background=C["panel"], foreground=C["fg"],
                    font=("Segoe UI", 9))
        s.configure("Dim.TLabel", background=C["panel"], foreground=C["fg_dim"],
                    font=("Segoe UI", 8))

        # LabelFrame
        s.configure("TLabelframe",
            background=C["bg"],
            bordercolor=NEU_DARK,
            darkcolor=NEU_DARK,
            lightcolor=NEU_LIGHT,
            relief="raised",
            borderwidth=4,
        )
        s.configure("TLabelframe.Label",
            background=C["bg"],
            foreground=C["accent"],
            font=("Segoe UI", 9, "bold"),
        )

        # Entry (sunken — 눌린 느낌)
        s.configure("TEntry",
            fieldbackground=C["input"],
            foreground=C["fg"],
            bordercolor=NEU_DARK,
            lightcolor=NEU_DARK,
            darkcolor=NEU_LIGHT,
            insertcolor=C["fg"],
            relief="sunken",
            borderwidth=3,
            padding=(4, 3),
        )
        s.map("TEntry",
            fieldbackground=[("readonly", C["panel"])],
            lightcolor=[("focus", C["accent"])],
            darkcolor=[("focus", NEU_LIGHT)],
        )

        # Checkbutton
        s.configure("TCheckbutton",
            background=C["bg"],
            foreground=C["fg"],
            focuscolor=C["bg"],
            font=("Segoe UI", 9),
        )
        s.map("TCheckbutton",
            background=[("active", C["bg"])],
            foreground=[("active", C["fg_bright"])],
        )

        # Button (기본 neumorphic — raised → sunken on press)
        s.configure("TButton",
            background=C["bg"],
            foreground=C["fg"],
            bordercolor=C["bg"],
            darkcolor=NEU_DARK,
            lightcolor=NEU_LIGHT,
            relief="raised",
            borderwidth=4,
            padding=(10, 5),
            font=("Segoe UI", 9),
        )
        s.map("TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", NEU_LIGHT), ("active", NEU_DARK)],
            lightcolor=[("pressed", NEU_DARK), ("active", NEU_LIGHT)],
            foreground=[("active", C["fg_bright"])],
        )

        # 시작 버튼 (teal)
        s.configure("Start.TButton",
            background=C["start"],
            foreground="#ffffff",
            darkcolor="#014d4d",
            lightcolor="#03dac6",
            relief="raised",
            borderwidth=4,
            font=("Segoe UI", 9, "bold"),
        )
        s.map("Start.TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", "#03dac6")],
            lightcolor=[("pressed", "#014d4d")],
            background=[("active", C["start_hl"]), ("pressed", C["start"])],
        )

        # 중지 버튼 (error red)
        s.configure("Stop.TButton",
            background=C["stop"],
            foreground="#ffffff",
            darkcolor="#6b0010",
            lightcolor="#cf6679",
            relief="raised",
            borderwidth=4,
            font=("Segoe UI", 9, "bold"),
        )
        s.map("Stop.TButton",
            relief=[("pressed", "sunken")],
            darkcolor=[("pressed", "#cf6679")],
            lightcolor=[("pressed", "#6b0010")],
            background=[("active", C["stop_hl"]), ("pressed", C["stop"])],
        )

        # Scrollbar
        s.configure("TScrollbar",
            background=C["bg"],
            troughcolor=C["input"],
            bordercolor=C["bg"],
            darkcolor=NEU_DARK,
            lightcolor=NEU_LIGHT,
            arrowcolor=C["fg_dim"],
            relief="raised",
            borderwidth=2,
        )
        s.map("TScrollbar",
            background=[("active", C["panel2"])],
        )

        # Separator
        s.configure("TSeparator", background=NEU_DARK)

        # Treeview
        s.configure("Treeview",
            background=C["input"],
            foreground=C["fg"],
            fieldbackground=C["input"],
            bordercolor=NEU_DARK,
            rowheight=26,
            font=("Segoe UI", 9),
        )
        s.configure("Treeview.Heading",
            background=C["bg"],
            foreground=C["fg_dim"],
            bordercolor=NEU_DARK,
            darkcolor=NEU_DARK,
            lightcolor=NEU_LIGHT,
            font=("Segoe UI", 9, "bold"),
            relief="raised",
            borderwidth=2,
        )
        s.map("Treeview",
            background=[("selected", C["sel"])],
            foreground=[("selected", C["fg_bright"])],
        )
        s.map("Treeview.Heading",
            background=[("active", C["panel2"])],
        )

    # ------------------------------------------------------------------
    # UI 구성
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # 카드 헬퍼
    # ------------------------------------------------------------------

    def _make_card(self, parent: tk.Widget, title: str = "") -> tk.Frame:
        """
        Neumorphic 카드: Card.TFrame (raised, NEU_LIGHT/NEU_DARK) → pad → body.
        반환값(body)의 ._outer 속성으로 외부에서 pack 배치.
        """
        outer = ttk.Frame(parent, style="Card.TFrame", padding=3)
        pad = tk.Frame(outer, bg=C["bg"])
        pad.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
        if title:
            tk.Label(pad, text=title.upper(),
                     bg=C["bg"], fg=C["fg_dim"],
                     font=("Segoe UI", 7, "bold")).pack(anchor=tk.W, pady=(0, 8))
            body = tk.Frame(pad, bg=C["bg"])
            body.pack(fill=tk.BOTH, expand=True)
        else:
            body = pad
        body._outer = outer
        return body

    # ------------------------------------------------------------------
    # UI 구성
    # ------------------------------------------------------------------

    def _build_ui(self, cfg: dict) -> None:
        G = 8   # card gap

        # ── 상태바 (하단 고정) ─────────────────────────────────────
        self._status_var = tk.StringVar(value="대기 중")
        tk.Label(self, textvariable=self._status_var,
                 bg=C["panel2"], fg=C["fg_dim"],
                 font=("Segoe UI", 8), anchor=tk.W,
                 padx=10, pady=3).pack(fill=tk.X, side=tk.BOTTOM)

        # ── 메인 컨테이너 ─────────────────────────────────────────────
        main = tk.Frame(self, bg=C["bg"])
        main.pack(fill=tk.BOTH, expand=True, padx=G, pady=G)

        # ── 상단 행: 코드캐치(좌) + 설정·제어(우) ─────────────────────
        top_row = tk.Frame(main, bg=C["bg"])
        top_row.pack(fill=tk.X, pady=(0, G))

        # ─── 코드 캐치 카드 (좌) ──────────────────────────────────────
        ca = self._make_card(top_row)
        ca._outer.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, G))

        tk.Label(ca, text="캐치된 코드",
                 bg=C["panel"], fg=C["accent"],
                 font=("Segoe UI", 8, "bold")).pack(anchor=tk.W, pady=(0, 6))

        code_box = tk.Frame(ca, bg=C["code_bg"],
                            highlightbackground=C["border"],
                            highlightthickness=1)
        code_box.pack(fill=tk.X)

        self._latest_code_var = tk.StringVar(value="—")
        tk.Label(code_box, textvariable=self._latest_code_var,
                 font=("Consolas", 34, "bold"),
                 fg=C["yellow"], bg=C["code_bg"], pady=10).pack()

        self._latest_meta_var = tk.StringVar(value="대기 중...")
        tk.Label(code_box, textvariable=self._latest_meta_var,
                 font=("Segoe UI", 8), fg=C["fg_dim"],
                 bg=C["code_bg"], pady=3).pack()

        tk.Frame(ca, bg=C["border"], height=1).pack(fill=tk.X, pady=(10, 5))

        hist_row = tk.Frame(ca, bg=C["panel"])
        hist_row.pack(fill=tk.X)
        tk.Label(hist_row, text="이전:", font=("Segoe UI", 8),
                 fg=C["fg_dim"], bg=C["panel"]).pack(side=tk.LEFT)
        self._history_var = tk.StringVar(value="없음")
        tk.Label(hist_row, textvariable=self._history_var,
                 font=("Consolas", 8), fg=C["accent"], bg=C["panel"],
                 wraplength=190, justify=tk.LEFT).pack(side=tk.LEFT, padx=4)

        # ─── 우측 컬럼 (모니터링 설정 + 제어 버튼) ───────────────────
        right_col = tk.Frame(top_row, bg=C["bg"])
        right_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ─── 모니터링 설정 카드 ────────────────────────────────────────
        cb = self._make_card(right_col, "모니터링 설정")
        cb._outer.pack(fill=tk.X, pady=(0, G))
        cb.columnconfigure(1, weight=1)

        ttk.Label(cb, text="채팅방 이름", style="Panel.TLabel").grid(
            row=0, column=0, sticky=tk.W, pady=3)
        self._room_var = tk.StringVar(value=cfg["room_name"])
        ttk.Entry(cb, textvariable=self._room_var).grid(
            row=0, column=1, columnspan=2, sticky=tk.EW, padx=(8, 0), pady=3)

        ttk.Label(cb, text="폴링 간격", style="Panel.TLabel").grid(
            row=1, column=0, sticky=tk.W, pady=3)
        ivf = ttk.Frame(cb, style="Panel.TFrame")
        ivf.grid(row=1, column=1, sticky=tk.W, padx=(8, 0), pady=3)
        self._interval_var = tk.StringVar(value=str(cfg["poll_interval"]))
        ttk.Entry(ivf, textvariable=self._interval_var, width=6).pack(side=tk.LEFT)
        ttk.Label(ivf, text=" 초", style="Panel.TLabel").pack(side=tk.LEFT)

        ttk.Label(cb, text="발신자 필터", style="Panel.TLabel").grid(
            row=2, column=0, sticky=tk.W, pady=3)
        self._sender_var = tk.StringVar(value=cfg["watch_sender"])
        ttk.Entry(cb, textvariable=self._sender_var).grid(
            row=2, column=1, sticky=tk.EW, padx=(8, 0), pady=3)
        ttk.Label(cb, text="비워두면 전체", style="Dim.TLabel").grid(
            row=2, column=2, sticky=tk.W, padx=6)

        self._topmost_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(cb, text="항상 위 표시",
                        variable=self._topmost_var,
                        command=self._toggle_topmost).grid(
            row=3, column=0, columnspan=3, sticky=tk.W, pady=(6, 0))

        # ─── 제어 버튼 카드 ────────────────────────────────────────────
        cc = self._make_card(right_col)
        cc._outer.pack(fill=tk.X)

        bf = tk.Frame(cc, bg=C["panel"])
        bf.pack(fill=tk.X)

        self._start_btn = ttk.Button(bf, text="▶  모니터링 시작",
                                     style="Start.TButton",
                                     command=self._start_monitor)
        self._start_btn.pack(side=tk.LEFT, padx=(0, 6))

        self._stop_btn = ttk.Button(bf, text="■  중지",
                                    style="Stop.TButton",
                                    command=lambda: self._stop_monitor(manual=True),
                                    state=tk.DISABLED)
        self._stop_btn.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Separator(bf, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=(0, 8), pady=2)

        ttk.Button(bf, text="로그 지우기",
                   command=self._clear_log).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(bf, text="텍스트 저장",
                   command=self._export_text).pack(side=tk.LEFT)

        # 코드 수동 입력
        bf2 = tk.Frame(cc, bg=C["panel"])
        bf2.pack(fill=tk.X, pady=(8, 0))
        self._manual_code_var = tk.StringVar()
        tk.Entry(bf2, textvariable=self._manual_code_var,
                 font=("Consolas", 11), fg=C["yellow"], bg=C["input"],
                 insertbackground=C["yellow"], relief=tk.FLAT,
                 width=12).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(bf2, text="코드 입력",
                   command=self._submit_manual_code).pack(side=tk.LEFT)

        # 예약 시작·종료
        tk.Frame(cc, bg=C["border"], height=1).pack(fill=tk.X, pady=(10, 6))
        tk.Label(cc, text="예약 시작·종료 시각  (HH:MM)",
                 font=("Segoe UI", 8), fg=C["fg_dim"],
                 bg=C["panel"]).pack(anchor=tk.W, pady=(0, 3))

        saved_scheds = cfg.get("monitor_schedules", DEFAULT_CONFIG["monitor_schedules"])
        while len(saved_scheds) < 2:
            saved_scheds.append({"enabled": False, "start": "", "stop": ""})

        def _fmt_time(var):
            t = var.get().strip().replace(":", "")
            if len(t) == 3:
                t = "0" + t
            if len(t) == 4 and t.isdigit():
                h, m = int(t[:2]), int(t[2:])
                if 0 <= h <= 23 and 0 <= m <= 59:
                    var.set(f"{h:02d}:{m:02d}")

        def _bind_time(entry, var):
            entry.bind("<FocusOut>", lambda *_: _fmt_time(var))
            entry.bind("<Return>",   lambda *_: _fmt_time(var))

        for i, sc in enumerate(saved_scheds[:2]):
            ev  = tk.BooleanVar(value=bool(sc.get("enabled", False)))
            sv  = tk.StringVar(value=str(sc.get("start", "")))
            stv = tk.StringVar(value=str(sc.get("stop", "")))
            self._monitor_schedules.append((ev, sv, stv))

            srow = tk.Frame(cc, bg=C["panel"])
            srow.pack(fill=tk.X, pady=1)
            ttk.Checkbutton(srow, variable=ev, style="TCheckbutton").pack(side=tk.LEFT)
            tk.Label(srow, text="시작", font=("Segoe UI", 8),
                     fg=C["fg_dim"], bg=C["panel"]).pack(side=tk.LEFT, padx=(2, 2))
            e_start = tk.Entry(srow, textvariable=sv,
                     font=("Consolas", 10), fg=C["ok"], bg=C["input"],
                     insertbackground=C["ok"], relief=tk.FLAT,
                     justify=tk.CENTER, width=6)
            e_start.pack(side=tk.LEFT)
            _bind_time(e_start, sv)
            tk.Label(srow, text="  종료", font=("Segoe UI", 8),
                     fg=C["fg_dim"], bg=C["panel"]).pack(side=tk.LEFT, padx=(6, 2))
            e_stop = tk.Entry(srow, textvariable=stv,
                     font=("Consolas", 10), fg=C["error"], bg=C["input"],
                     insertbackground=C["error"], relief=tk.FLAT,
                     justify=tk.CENTER, width=6)
            e_stop.pack(side=tk.LEFT)
            _bind_time(e_stop, stv)

        # ─── 자동 입력 설정 카드 ──────────────────────────────────────
        cd = self._make_card(main, "자동 입력 설정")
        cd._outer.pack(fill=tk.X, pady=(0, G))
        cd.columnconfigure(1, weight=1)
        cd.columnconfigure(3, weight=1)

        self._auto_var = tk.BooleanVar(value=cfg.get("auto_input", False))
        auto_chk = ttk.Checkbutton(cd, text="코드 캐치 시 자동으로 사이트에 입력",
                                   variable=self._auto_var)
        auto_chk.grid(row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 6))

        wf = ttk.Frame(cd, style="Panel.TFrame")
        wf.grid(row=0, column=3, sticky=tk.W, padx=(8, 0), pady=(0, 6))
        ttk.Label(wf, text="워커 수", foreground=C["accent"],
                  style="Panel.TLabel").pack(side=tk.LEFT)
        self._workers_var = tk.StringVar(value=str(cfg.get("workers", 2)))
        tk.Spinbox(wf, from_=1, to=30, textvariable=self._workers_var,
                   width=3, font=("Segoe UI", 9),
                   bg=C["input"], fg=C["accent"], buttonbackground=C["bg"],
                   relief=tk.FLAT).pack(side=tk.LEFT, padx=(4, 2))
        ttk.Label(wf, text="개", foreground=C["accent"],
                  style="Panel.TLabel").pack(side=tk.LEFT)

        if not API_OK and not SELENIUM_OK:
            ttk.Label(cd, text="⚠  curl_cffi/selenium 미설치 — 자동 입력 불가",
                      foreground=C["error"],
                      style="Panel.TLabel").grid(
                row=0, column=4, sticky=tk.W, padx=(16, 0))
            self._auto_var.set(False)
            auto_chk.config(state=tk.DISABLED)

        saved_urls = cfg.get("site_urls", DEFAULT_CONFIG["site_urls"])
        if isinstance(saved_urls, str):
            saved_urls = [saved_urls] + [""] * 9
        while len(saved_urls) < 10:
            saved_urls.append("")
        if self._shared_site_url_vars is not None:
            self._site_url_vars = self._shared_site_url_vars
        else:
            self._site_url_vars = [tk.StringVar(value=saved_urls[i]) for i in range(10)]
        self._site_url_entries = []

        for pr, (li, ri) in enumerate([(0,1),(2,3),(4,5),(6,7),(8,9)], start=1):
            ttk.Label(cd, text=f"사이트 {li+1}", style="Panel.TLabel").grid(
                row=pr, column=0, sticky=tk.W, pady=2)
            e_l = tk.Entry(cd, textvariable=self._site_url_vars[li],
                           bg=C["input"], fg=C["fg"], insertbackground=C["fg"],
                           relief=tk.FLAT, bd=1)
            e_l.grid(row=pr, column=1, sticky=tk.EW, padx=(6, 14), pady=2)
            self._site_url_entries.append(e_l)
            ttk.Label(cd, text=f"사이트 {ri+1}", style="Panel.TLabel").grid(
                row=pr, column=2, sticky=tk.W, pady=2)
            e_r = tk.Entry(cd, textvariable=self._site_url_vars[ri],
                           bg=C["input"], fg=C["fg"], insertbackground=C["fg"],
                           relief=tk.FLAT, bd=1)
            e_r.grid(row=pr, column=3, sticky=tk.EW, padx=(6, 0), pady=2)
            self._site_url_entries.append(e_r)

        ttk.Label(cd, text="Chrome 포트", style="Panel.TLabel").grid(
            row=6, column=0, sticky=tk.W, pady=(6, 2))
        pf = ttk.Frame(cd, style="Panel.TFrame")
        pf.grid(row=6, column=1, sticky=tk.W, padx=(6, 14), pady=(6, 2))
        self._port_var = tk.StringVar(value=str(cfg.get("chrome_port", 9222)))
        ttk.Entry(pf, textvariable=self._port_var, width=7).pack(side=tk.LEFT)
        ttk.Label(pf, text="  기본 9222", style="Dim.TLabel").pack(side=tk.LEFT)
        ttk.Button(cd, text="접속 테스트", command=self._test_connections).grid(
            row=6, column=2, columnspan=2, sticky=tk.EW, padx=(6, 0), pady=(6, 2))

        ttk.Label(cd, text="현재 계정", style="Panel.TLabel").grid(
            row=7, column=0, sticky=tk.W, pady=(8, 2))
        self._current_acct_var = tk.StringVar()
        ttk.Label(cd, textvariable=self._current_acct_var,
                  style="Panel.TLabel",
                  foreground=C["accent"]).grid(
            row=7, column=1, columnspan=2, sticky=tk.W, padx=(6, 0), pady=(8, 2))
        ttk.Button(cd, text="계정 관리",
                   command=self._open_account_manager).grid(
            row=7, column=3, sticky=tk.E, pady=(8, 2))

        ttk.Label(cd, text="프록시 IP", style="Panel.TLabel").grid(
            row=8, column=0, sticky=tk.W, pady=(4, 2))
        self._proxy_info_var = tk.StringVar()
        self._update_proxy_label()
        ttk.Label(cd, textvariable=self._proxy_info_var,
                  style="Panel.TLabel",
                  foreground=C["accent"]).grid(
            row=8, column=1, columnspan=2, sticky=tk.W, padx=(6, 0), pady=(4, 2))
        ttk.Button(cd, text="IP 관리",
                   command=self._open_proxy_manager).grid(
            row=8, column=3, sticky=tk.E, pady=(4, 2))
        self._update_current_acct_label()


        # ─── 메시지 로그 카드 ──────────────────────────────────────────
        ce = self._make_card(main, "메시지 로그")
        ce._outer.pack(fill=tk.BOTH, expand=True)

        self._log = tk.Text(
            ce, state=tk.DISABLED,
            background=C["log_bg"], foreground=C["fg"],
            insertbackground=C["fg"],
            relief=tk.FLAT, wrap=tk.WORD,
            font=("Consolas", 10),
            selectbackground=C["sel"],
            padx=6, pady=4,
        )
        sb = ttk.Scrollbar(ce, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._log.pack(fill=tk.BOTH, expand=True)

        self._log.tag_configure("date_time",   foreground=C["fg_dim"])
        self._log.tag_configure("timestamp",   foreground="#2471a3")
        self._log.tag_configure("sender",      foreground="#1a7a5e",
                                font=("Consolas", 10, "bold"))
        self._log.tag_configure("system",      foreground=C["system"],
                                font=("Consolas", 10, "italic"))
        self._log.tag_configure("error",       foreground=C["error"])
        self._log.tag_configure("caught_code", foreground=C["yellow"],
                                font=("Consolas", 10, "bold"))
        self._log.tag_configure("auto_ok",     foreground=C["ok"],
                                font=("Consolas", 10, "italic"))

    # ------------------------------------------------------------------
    # 설정 헬퍼
    # ------------------------------------------------------------------

    def _get_current_config(self) -> dict:
        try:
            poll_interval = float(self._interval_var.get())
        except ValueError:
            poll_interval = DEFAULT_CONFIG["poll_interval"]
        try:
            chrome_port = int(self._port_var.get())
        except ValueError:
            chrome_port = 9222
        try:
            workers = max(1, min(30, int(self._workers_var.get())))
        except ValueError:
            workers = 2
        return {
            "room_name": self._room_var.get().strip(),
            "poll_interval": poll_interval,
            "watch_sender": self._sender_var.get().strip(),
            "auto_input": self._auto_var.get(),
            "chrome_port": chrome_port,
            "workers": workers,
            "site_urls": [v.get().strip() for v in self._site_url_vars],
            "accounts": self._accounts,
            "account_index": self._account_idx,
            "proxies": self._proxies,
            "app_password": self._app_password,
            "monitor_schedules": [
                {"enabled": bool(ev.get()),
                 "start": sv.get().strip(),
                 "stop": stv.get().strip()}
                for ev, sv, stv in self._monitor_schedules
            ],
        }

    # ------------------------------------------------------------------
    # 계정 관리
    # ------------------------------------------------------------------

    def _update_current_acct_label(self) -> None:
        if not self._accounts:
            self._current_acct_var.set("등록된 계정 없음 — [계정 관리]에서 추가하세요")
        else:
            enabled = [a for a in self._accounts if a.get("enabled", True)]
            total = len(self._accounts)
            self._current_acct_var.set(
                f"총 {total}개 계정 등록 | 활성 {len(enabled)}개")

    def _test_connections(self) -> None:
        """입력된 사이트 URL 접속 가능 여부를 테스트. 실패 URL은 빨간색으로 표시."""
        import urllib.request
        pairs = [(self._site_url_entries[i], v.get().strip())
                 for i, v in enumerate(self._site_url_vars) if v.get().strip()]
        if not pairs:
            messagebox.showinfo("접속 테스트", "입력된 사이트 URL이 없습니다.")
            return

        # 테스트 전 모든 대상 배경 초기화
        for entry, _ in pairs:
            entry.config(bg=C["input"])

        def _run():
            lines = []
            for entry, url in pairs:
                try:
                    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                    with urllib.request.urlopen(req, timeout=8):
                        lines.append(f"✓  {url}")
                        self.after(0, lambda e=entry: e.config(bg=C["input"]))
                except Exception as ex:
                    lines.append(f"✗  {url}\n     {ex}")
                    self.after(0, lambda e=entry: e.config(bg="#5c1a1a"))
            self.after(0, lambda m="\n\n".join(lines): messagebox.showinfo("접속 테스트 결과", m))

        messagebox.showinfo("접속 테스트", f"{len(pairs)}개 사이트 접속 테스트 중...\n완료되면 결과 팝업이 표시됩니다.")
        threading.Thread(target=_run, daemon=True).start()

    def _update_proxy_label(self) -> None:
        n = len([p for p in self._proxies if p.strip()])
        if n == 0:
            self._proxy_info_var.set("등록된 IP 없음 — 기본 IP로 접속")
        else:
            self._proxy_info_var.set(f"총 {n}개 IP 등록 | 워커별 자동 배분")

    def _open_proxy_manager(self) -> None:
        def on_save(proxies):
            self._proxies = proxies
            self._update_proxy_label()
            save_config(self._get_current_config())
        ProxyManagerDialog(self, self._proxies, on_save)

    def _open_account_manager(self) -> None:
        if self._app_password:
            pw = tk.simpledialog.askstring("비밀번호", "비밀번호를 입력하세요:", show="*")
            if pw != self._app_password:
                if pw is not None:
                    messagebox.showerror("오류", "비밀번호가 틀립니다.")
                return

        def on_save(accounts, current_idx):
            self._accounts = accounts
            self._account_idx = current_idx
            self._update_current_acct_label()
            save_config(self._get_current_config())

        AccountManagerDialog(self, self._accounts, self._account_idx, on_save)

    def _submit_manual_code(self) -> None:
        code = self._manual_code_var.get().strip()
        if not code:
            self._append_log("코드를 입력하세요.", tag="error")
            return
        if not CODE_PATTERN.match(code):
            self._append_log("9자리 영숫자 코드를 입력하세요.", tag="error")
            return
        self._manual_code_var.set("")
        self._append_log(f"📝 수동 코드 입력: {code}", tag="system")
        self._update_catch_panel(code, "수동", "직접입력")
        if self._auto_var.get():
            self._start_auto_input(code)

    def _advance_account(self) -> None:
        if self._accounts:
            self._account_idx = (self._account_idx + 1) % len(self._accounts)
            self.after(0, self._update_current_acct_label)
            save_config(self._get_current_config())


    # ------------------------------------------------------------------
    # 제어
    # ------------------------------------------------------------------

    def _start_monitor(self) -> None:
        room_name = self._room_var.get().strip()
        if not room_name:
            self._append_log("채팅방 이름을 입력하세요.", tag="error")
            return
        try:
            poll_interval = float(self._interval_var.get())
        except ValueError:
            self._append_log("폴링 간격은 숫자로 입력하세요.", tag="error")
            return

        save_config(self._get_current_config())

        self._stop_event = threading.Event()
        self._auto_cancel_event.clear()
        self._monitor_thread = threading.Thread(
            target=run_monitor,
            kwargs={
                "room_name": room_name,
                "poll_interval": poll_interval,
                "watch_sender": self._sender_var.get().strip(),
                "on_new_message": self._on_new_message,
                "stop_event": self._stop_event,
            },
            daemon=True,
        )
        self._monitor_thread.start()

        self._start_btn.config(state=tk.DISABLED)
        self._stop_btn.config(state=tk.NORMAL)
        self._status_var.set(f"모니터링 중  |  채팅방: {room_name}")
        self._append_log(f"── 모니터링 시작: '{room_name}' ──", tag="system")

    def _stop_monitor(self, manual: bool = False) -> None:
        if self._stop_event:
            self._stop_event.set()
        self._auto_cancel_event.set()
        self._processed_codes.clear()
        # 자동입력 진행 중이면 chromedriver 강제 종료 (즉시 중단)
        if self._auto_input_active:
            try:
                import subprocess
                subprocess.run(["taskkill", "/f", "/im", "chromedriver.exe"],
                             capture_output=True, timeout=5)
                subprocess.run(["taskkill", "/f", "/im", "undetected_chromedriver.exe"],
                             capture_output=True, timeout=5)
                self._append_log("── chromedriver 강제 종료 ──", tag="system")
            except Exception as e:
                logger.warning(f"chromedriver 종료 실패: {e}")
        if manual:
            self._code_detected_this_slot = True  # 수동 정지 시 스케줄러 재시작 방지
        self._start_btn.config(state=tk.NORMAL)
        self._stop_btn.config(state=tk.DISABLED)
        self._status_var.set("중지됨")
        self._append_log("── 모니터링 중지 ──", tag="system")

    # ------------------------------------------------------------------
    # 예약 시작·종료 스케줄러
    # ------------------------------------------------------------------

    def _start_monitor_scheduler(self) -> None:
        self._last_sched_action = ["", ""]
        self._last_sched_log_min = ""
        for i, (ev, sv, stv) in enumerate(self._monitor_schedules):
            logger.info(f"스케줄 초기화 [{i}]: enabled={ev.get()}, start={sv.get()}, stop={stv.get()}")
        self._check_monitor_schedule()

    def _check_monitor_schedule(self) -> None:
        """독립 루프: 10초마다 스케줄 체크."""
        try:
            now_dt = datetime.datetime.now()
            today  = now_dt.strftime("%Y-%m-%d")
            now    = now_dt.strftime("%H:%M")
            is_running = (self._stop_event is not None and
                          not self._stop_event.is_set())

            def _norm(t: str) -> str:
                t = t.strip()
                if not t:
                    return ""
                try:
                    return datetime.datetime.strptime(t, "%H:%M").strftime("%H:%M")
                except ValueError:
                    try:
                        return datetime.datetime.strptime(t, "%I:%M").strftime("%H:%M")
                    except ValueError:
                        return t

            def _in_range(start: str, stop: str, t: str) -> bool:
                """시간 범위 체크 (자정 넘김 지원)."""
                if not start or not stop:
                    return False
                if start <= stop:
                    return start <= t < stop
                else:
                    # 자정 넘김: 예) 23:30~00:00 → 23:30~23:59 또는 00:00
                    return t >= start or t < stop

            def _near_schedule(start: str, stop: str, now_min: int) -> bool:
                """스케줄 시간 전후 5분 이내인지 체크."""
                if not start or not stop:
                    return False
                try:
                    sh, sm = map(int, start.split(":"))
                    eh, em = map(int, stop.split(":"))
                    start_min = sh * 60 + sm
                    stop_min  = eh * 60 + em
                    # 시작 5분 전 ~ 종료 5분 후
                    ds = (now_min - start_min) % 1440
                    de = (stop_min - now_min) % 1440
                    return ds <= 5 or ds >= 1435 or de <= 5 or de >= 1435
                except ValueError:
                    return True  # 파싱 실패 시 항상 로그

            now_minutes = now_dt.hour * 60 + now_dt.minute
            near_any = False

            for i, (ev, sv, stv) in enumerate(self._monitor_schedules):
                if not ev.get():
                    continue
                start_t = _norm(sv.get())
                stop_t  = _norm(stv.get())

                if _near_schedule(start_t, stop_t, now_minutes):
                    near_any = True

                in_range = _in_range(start_t, stop_t, now)

                # 범위 안에 있으면 모니터링 자동 시작 (자동입력 중이거나 코드 감지 후엔 대기)
                if in_range and not is_running and not self._auto_input_active and not self._code_detected_this_slot:
                    logger.debug(f"스케줄[{i}]: 시작 조건 충족 — code_detected={self._code_detected_this_slot}, auto_input={self._auto_input_active}")
                    key = f"start-{today}-{now}-{i}"
                    if self._last_sched_action[i] != key:
                        self._last_sched_action[i] = key
                        logger.info(f"스케줄 자동 시작: {now} (범위 {start_t}~{stop_t})")
                        self._append_log(f"⏰ 예약 시작: {now}", tag="system")
                        self._start_monitor()
                        is_running = True
                elif in_range and is_running:
                    pass  # 이미 실행 중 — 정상
                elif in_range and self._auto_input_active:
                    if now != self._last_sched_log_min:
                        logger.debug(f"스케줄[{i}]: 범위 내지만 자동입력 중 — 대기")
                elif in_range and not is_running and self._code_detected_this_slot:
                    if now != self._last_sched_log_min:
                        logger.debug(f"스케줄[{i}]: 코드 감지 슬롯 — 재시작 차단 중")

                if stop_t and stop_t == now and is_running:
                    key = f"stop-{today}-{now}-{i}"
                    if self._last_sched_action[i] != key:
                        self._last_sched_action[i] = key
                        logger.info(f"스케줄 자동 종료: {now}")
                        self._append_log(f"⏰ 예약 종료: {now}", tag="system")
                        self._stop_monitor()
                        is_running = False

                # 범위 밖이면 코드 감지 플래그 리셋
                if not in_range:
                    self._code_detected_this_slot = False

            # 스케줄 전후 5분 이내일 때만 로그 출력
            if near_any and now != self._last_sched_log_min:
                self._last_sched_log_min = now
                logger.debug(f"스케줄러 체크: {now}, 모니터링={'실행중' if is_running else '대기'}")
        except Exception as e:
            logger.error(f"스케줄러 오류: {e}")
        finally:
            self.after(10000, self._check_monitor_schedule)

    # ------------------------------------------------------------------
    # 에이전트 연동 독립 루프 (5초마다)
    # ------------------------------------------------------------------

    def _start_agent_sync(self) -> None:
        """에이전트 연동 루프 시작."""
        self._agent_sync_loop()

    def _agent_sync_loop(self) -> None:
        """독립 루프: 5초마다 status.json 기록 + commands.json 감시 + config reload."""
        try:
            self._write_status_file()
            self._check_external_commands()
            self._check_config_reload()
        except Exception as e:
            logger.error(f"에이전트 연동 오류: {e}")
        finally:
            self.after(5000, self._agent_sync_loop)

    def _clear_log(self) -> None:
        self._log_lines.clear()
        self._log.config(state=tk.NORMAL)
        self._log.delete("1.0", tk.END)
        self._log.config(state=tk.DISABLED)

    def _toggle_topmost(self) -> None:
        self.attributes("-topmost", self._topmost_var.get())

    def _on_close(self) -> None:
        if self._stop_event:
            self._stop_event.set()
        self._save_log_data()
        save_config(self._get_current_config())
        # status.json 삭제 (오프라인 표시용)
        try:
            STATUS_PATH.unlink(missing_ok=True)
        except Exception:
            pass
        if self._win is not None:
            self._win.destroy()

    # ------------------------------------------------------------------
    # 에이전트 연동 — status.json / commands.json / config reload
    # ------------------------------------------------------------------

    def _write_status_file(self) -> None:
        """에이전트가 읽어 서버에 보고할 상태 파일 기록"""
        try:
            is_running = (self._stop_event is not None and
                          not self._stop_event.is_set())
            enabled_count = sum(1 for a in self._accounts if a.get("enabled", True))
            # 최근 로그 20줄
            recent_logs = [l.get("text", "") for l in self._log_lines[-50:]]
            data = {
                "monitoring_active": is_running,
                "auto_input_active": self._auto_input_active,
                "last_code": self._last_code_for_status,
                "last_code_time": self._last_code_time_for_status,
                "success_count": self._success_count,
                "fail_count": self._fail_count,
                "app_version": getattr(self, '_override_version', None) or (VERSION_SELENIUM if _FORCE_SELENIUM else VERSION),
                "account_count": len(self._accounts),
                "enabled_account_count": enabled_count,
                "recent_logs": recent_logs,
                "recent_errors": self._recent_errors[-10:],
            }
            with open(STATUS_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
        except Exception as e:
            logger.debug(f"status.json 기록 실패: {e}")

    def _check_external_commands(self) -> None:
        """에이전트가 기록한 commands.json 읽고 실행"""
        if not COMMANDS_PATH.exists():
            return
        try:
            with open(COMMANDS_PATH, encoding="utf-8") as f:
                commands = json.load(f)
            if not commands:
                return
            # 파일 즉시 비우기
            with open(COMMANDS_PATH, "w", encoding="utf-8") as f:
                json.dump([], f)

            for cmd in commands:
                cmd_type = cmd.get("type", "")
                payload = cmd.get("payload", {})
                logger.info(f"외부 명령 수신: {cmd_type}")

                if cmd_type == "start_monitor":
                    is_running = (self._stop_event is not None and
                                  not self._stop_event.is_set())
                    if not is_running and not self._code_detected_this_slot:
                        self._start_monitor()
                    elif self._code_detected_this_slot:
                        logger.info("코드 감지 슬롯 — 외부 start_monitor 무시")
                elif cmd_type == "stop_monitor":
                    self._stop_monitor(manual=True)
                elif cmd_type == "submit_code":
                    code = payload.get("code", "")
                    source = payload.get("source", "auto")
                    if code and CODE_PATTERN.match(code):
                        # 원격 명령은 중복 체크 초기화 (대시보드 수동 재전송 허용)
                        self._processed_codes.discard(code)
                        if source == "manual":
                            self._append_log(f"📡 수동 전송 코드 수신: {code}", tag="system")
                        else:
                            self._append_log(f"📡 자동 감지 코드 수신: {code}", tag="system")
                        self._update_catch_panel(code, "원격", "서버")
        except Exception as e:
            logger.error(f"commands.json 처리 오류: {e}")

    def _check_config_reload(self) -> None:
        """에이전트가 config.json을 덮어쓴 경우 감지 → 리로드"""
        try:
            if not CONFIG_PATH.exists():
                return
            mtime = CONFIG_PATH.stat().st_mtime
            if self._last_config_mtime == 0:
                self._last_config_mtime = mtime
                return
            if mtime > self._last_config_mtime:
                self._last_config_mtime = mtime
                # 앱 자체 저장인 경우 무시
                if mtime == _last_self_save_mtime:
                    return
                logger.info("config.json 외부 변경 감지 — 리로드")
                cfg = load_config()
                self._accounts = list(cfg.get("accounts", []))
                self._proxies = list(cfg.get("proxies", []))
                self._app_password = cfg.get("app_password", "")
                self._room_var.set(cfg.get("room_name", ""))
                self._interval_var.set(str(cfg.get("poll_interval", 3)))
                self._sender_var.set(cfg.get("watch_sender", ""))
                self._auto_var.set(cfg.get("auto_input", False))
                self._workers_var.set(str(cfg.get("workers", 2)))
                self._port_var.set(str(cfg.get("chrome_port", 9222)))
                # 사이트 URL 업데이트
                site_urls = cfg.get("site_urls", [])
                for i, sv in enumerate(self._site_url_vars):
                    sv.set(site_urls[i] if i < len(site_urls) else "")
                # 스케줄 업데이트
                scheds = cfg.get("monitor_schedules", [])
                for i, (ev, sv, stv) in enumerate(self._monitor_schedules):
                    if i < len(scheds):
                        ev.set(scheds[i].get("enabled", False))
                        sv.set(scheds[i].get("start", ""))
                        stv.set(scheds[i].get("stop", ""))
                self._append_log("📡 설정이 서버에서 업데이트되었습니다", tag="system")
        except Exception as e:
            logger.debug(f"config reload 체크 오류: {e}")

    # ------------------------------------------------------------------
    # 스레드 → GUI
    # ------------------------------------------------------------------

    def _on_new_message(self, msg: ChatMessage) -> None:
        self._msg_queue.put(msg)

    def _poll_queue(self) -> None:
        try:
            try:
                while True:
                    item = self._msg_queue.get_nowait()
                    if isinstance(item, ChatMessage):
                        self._handle_message(item)
                    elif isinstance(item, tuple):
                        tag, text = item
                        if tag == "__auto_start__":
                            self._auto_input_active = True
                            self._stop_btn.config(state=tk.NORMAL)
                        elif tag == "__auto_end__":
                            self._auto_input_active = False
                            self._code_detected_this_slot = True
                            # 모니터도 중지된 경우에만 stop 버튼 비활성화
                            if self._stop_event is None or self._stop_event.is_set():
                                self._stop_btn.config(state=tk.DISABLED)
                        else:
                            self._append_log(text, tag=tag)
                    elif isinstance(item, str):
                        self._append_log(item, tag="system")
            except queue.Empty:
                pass

            # 모니터 스레드가 종료됐고 자동입력도 없을 때 버튼 상태 복구
            if (self._monitor_thread is not None
                    and not self._monitor_thread.is_alive()
                    and self._stop_btn["state"] == tk.NORMAL
                    and not self._auto_input_active):
                self._start_btn.config(state=tk.NORMAL)
                self._stop_btn.config(state=tk.DISABLED)
                self._status_var.set("모니터링 종료")
                self._monitor_thread = None

        except Exception as e:
            logger.error(f"_poll_queue 오류: {e}")
        finally:
            self.after(200, self._poll_queue)

    # ------------------------------------------------------------------
    # 메시지 처리
    # ------------------------------------------------------------------

    def _handle_message(self, msg: ChatMessage) -> None:
        code = msg.content.strip()
        is_code = bool(CODE_PATTERN.match(code))
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 로그 영구 저장용 plain 라인
        if is_code:
            plain = f"[{now_str}] [{msg.timestamp_str}] {msg.sender}: {code}  ★"
            self._log_lines.append({"text": plain, "tag": "caught_code"})
        else:
            plain = f"[{now_str}] [{msg.timestamp_str}] {msg.sender}: {msg.content}"
            self._log_lines.append({"text": plain, "tag": ""})
        if len(self._log_lines) > 10000:
            self._log_lines = self._log_lines[-10000:]

        self._log.config(state=tk.NORMAL)
        self._log.insert(tk.END, f"[{now_str}] ", "date_time")
        self._log.insert(tk.END, f"[{msg.timestamp_str}] ", "timestamp")
        self._log.insert(tk.END, f"{msg.sender}", "sender")
        self._log.insert(tk.END, ": ")
        if is_code:
            self._log.insert(tk.END, f"{code}  ★\n", "caught_code")
        else:
            self._log.insert(tk.END, f"{msg.content}\n")
        self._log.see(tk.END)
        self._log.config(state=tk.DISABLED)

        if is_code:
            # 코드 감지 → 항상 플래그 설정 (모니터링 재시작 방지)
            self._code_detected_this_slot = True
            self._update_catch_panel(code, msg.timestamp_str, msg.sender)
            # 코드 추출 성공 → 카카오톡 모니터 즉시 중단
            if self._stop_event and not self._stop_event.is_set():
                self._stop_event.set()
                self._start_btn.config(state=tk.NORMAL)
                # 자동입력 실행 중이면 중지 버튼 유지 (취소 가능하도록)
                if not self._auto_input_active:
                    self._stop_btn.config(state=tk.DISABLED)
                self._append_log("── 코드 감지로 모니터링 자동 중지 ──", tag="system")

    def _update_catch_panel(self, code: str, ts: str, sender: str) -> None:
        self._caught_codes_full.append({
            "code": code, "ts": ts, "sender": sender,
            "datetime": datetime.datetime.now().isoformat(timespec="seconds"),
        })
        # 메모리 누수 방지: 최대 1000개 유지
        if len(self._caught_codes_full) > 1000:
            self._caught_codes_full = self._caught_codes_full[-1000:]
        current = self._latest_code_var.get()
        if current != "—":
            self._caught_history.insert(0, current)
            self._caught_history = self._caught_history[:8]

        self._latest_code_var.set(code)
        self._latest_meta_var.set(f"{ts}  |  {sender}")
        self._history_var.set("  ".join(self._caught_history) if self._caught_history else "없음")
        self._status_var.set(f"코드 캐치: {code}  ({ts} / {sender})")

        if not API_OK and not SELENIUM_OK:
            self._append_log("⚠ curl_cffi/selenium 미설치 — 자동 입력 불가", tag="error")
            return
        if not self._auto_var.get():
            self._append_log("ℹ 자동 입력 비활성", tag="system")
            return
        if self._auto_input_active:
            self._append_log(f"자동입력 진행 중 — 코드 무시: {code}", tag="system")
            return
        if code in self._processed_codes:
            self._append_log(f"중복 코드 무시: {code}", tag="system")
            return
        self._processed_codes.add(code)

        try:
            port = int(self._port_var.get())
        except ValueError:
            port = 9222
        site_urls = [v.get().strip() for v in self._site_url_vars if v.get().strip()]
        if not site_urls:
            self._append_log("⚠ 사이트 주소를 입력하세요.", tag="error")
            return

        threading.Thread(
            target=self._run_auto_input,
            args=(code, port, site_urls),
            daemon=True,
        ).start()

    def _run_auto_input(self, code: str, port: int, site_urls: list) -> None:
        import concurrent.futures
        if self._auto_cancel_event.is_set():
            return

        # GUI에 자동입력 시작 알림 (중지 버튼 활성 유지)
        self._msg_queue.put(("__auto_start__", ""))

        try:
            accounts = [a for a in self._accounts if a.get("enabled", True)]
            if not accounts:
                self._msg_queue.put(("error", "⚠ 활성화된 계정이 없습니다. 계정 관리에서 ☑ 체크하세요."))
                return

            try:
                workers = max(1, min(30, int(self._workers_var.get())))
            except ValueError:
                workers = 2

            total = len(accounts)
            import datetime as _dt
            _now = _dt.datetime.now().strftime("%H:%M:%S")
            self._msg_queue.put(("system", f"━━ 자동 입력 시작: {code} | 활성 계정 {total}개 (워커 {workers}개 병렬) | {_now} ━━"))

            results = {}  # index → ("ok"|"fail", message)

            # 프록시 풀 → 워커에 round-robin 배분, 없으면 계정 자체 프록시 사용
            proxy_pool = [p.strip() for p in self._proxies if p.strip()]

            def _process(i: int, acct: dict) -> None:
                acct_label = f"[{i+1}/{total}] {acct['email']}"
                if proxy_pool:
                    proxy = proxy_pool[i % len(proxy_pool)]
                else:
                    proxy = _valid_proxy(acct.get("proxy", ""))
                if proxy:
                    self._msg_queue.put(("system", f"→ {acct_label} 처리 중... (IP: {proxy})"))
                else:
                    self._msg_queue.put(("system", f"→ {acct_label} 처리 중..."))

                def _status(msg: str) -> None:
                    self._msg_queue.put(("system", f"  · [{i+1}] {msg}"))

                def _log_err(stage: str, error_type: str, msg: str) -> None:
                    entry = {
                        "time": datetime.datetime.now().strftime("%H:%M:%S"),
                        "email": acct["email"], "code": code,
                        "stage": stage, "error_type": error_type,
                        "message": msg[:200], "result": "실패",
                    }
                    write_result(BASE_DIR, "픽", entry)
                    self._recent_errors.append(entry)
                    if len(self._recent_errors) > 30:
                        self._recent_errors = self._recent_errors[-30:]

                try:
                    submit_order_code(
                        code, port, site_urls,
                        email=acct["email"],
                        password=acct["password"],
                        status_cb=_status,
                        cancel_event=self._auto_cancel_event,
                        proxy=proxy,
                    )
                    results[i] = ("ok", acct_label)
                    self._msg_queue.put(("auto_ok", f"✓ 완료: {acct_label}"))
                    write_result(BASE_DIR, "픽", {
                        "time": datetime.datetime.now().strftime("%H:%M:%S"),
                        "email": acct["email"], "code": code,
                        "stage": "완료", "error_type": "", "message": "", "result": "성공",
                    })
                except AutoCancelled:
                    results[i] = ("cancel", acct_label)
                except LoginFailed as e:
                    results[i] = ("login_fail", acct_label)
                    self._msg_queue.put(("error", f"✗ 로그인 실패: {acct_label} — {e}"))
                    _log_err("로그인", "로그인 실패", str(e))
                except InvalidParameter as e:
                    results[i] = ("fail", acct_label)
                    self._msg_queue.put(("error", f"✗ 코드 무효: {acct_label} — {e}"))
                    _log_err("코드 입력", "코드 무효", str(e))
                except Exception as e:
                    stage, error_type = classify_exception(e)
                    results[i] = ("fail", acct_label)
                    self._msg_queue.put(("error", f"✗ 실패: {acct_label} [{error_type}] — {str(e)[:100]}"))
                    _log_err(stage, error_type, str(e))

            with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
                futures = []
                for i, acct in enumerate(accounts):
                    futures.append(pool.submit(_process, i, acct))
                    if i < len(accounts) - 1:  # 마지막 계정 제외하고 모두 간격 적용
                        import random
                        delay = random.uniform(*STAGGER_DELAY_API) if API_OK else STAGGER_DELAY
                        if self._auto_cancel_event.wait(delay):
                            break
                concurrent.futures.wait(futures)

            cancelled = any(v[0] == "cancel" for v in results.values())
            if cancelled:
                self._msg_queue.put(("system", "── 자동 입력 취소됨 ──"))
                self._code_detected_this_slot = True

            # ── 실패 계정 자동 재시도 (로그인 실패 + 일반 오류, 워커 병렬) ────────
            failed_indices = [i for i, v in results.items() if v[0] in ("login_fail", "fail")]
            if failed_indices and not cancelled and not self._auto_cancel_event.is_set():
                retry_total = len(failed_indices)
                _now_r = _dt.datetime.now().strftime("%H:%M:%S")
                self._msg_queue.put(("system", f"── 실패 {retry_total}개 재시도 시작 (워커 {workers}개) | {_now_r} ──"))

                retry_results = {}
                retry_lock = threading.Lock()

                def _retry_process(r_idx: int, orig_i: int) -> None:
                    acct = accounts[orig_i]
                    acct_label_r = f"[재시도 {r_idx+1}/{retry_total}] {acct['email']}"
                    if proxy_pool:
                        proxy_r = proxy_pool[orig_i % len(proxy_pool)]
                    else:
                        proxy_r = _valid_proxy(acct.get("proxy", ""))
                    self._msg_queue.put(("system", f"→ {acct_label_r} 처리 중..."))

                    def _status_r(msg: str) -> None:
                        self._msg_queue.put(("system", f"  · [{r_idx+1}] {msg}"))

                    def _log_err_r(stage: str, error_type: str, msg: str) -> None:
                        entry = {
                            "time": datetime.datetime.now().strftime("%H:%M:%S"),
                            "email": acct["email"], "code": code,
                            "stage": stage, "error_type": error_type,
                            "message": msg[:200], "result": "재시도실패",
                        }
                        write_result(BASE_DIR, "픽", entry)
                        self._recent_errors.append(entry)
                        if len(self._recent_errors) > 30:
                            self._recent_errors = self._recent_errors[-30:]

                    try:
                        submit_order_code(
                            code, port, site_urls,
                            email=acct["email"],
                            password=acct["password"],
                            status_cb=_status_r,
                            cancel_event=self._auto_cancel_event,
                            proxy=proxy_r,
                        )
                        with retry_lock:
                            retry_results[r_idx] = ("ok", acct_label_r)
                        self._msg_queue.put(("auto_ok", f"✓ 완료: {acct_label_r}"))
                    except AutoCancelled:
                        with retry_lock:
                            retry_results[r_idx] = ("cancel", acct_label_r)
                    except LoginFailed as e:
                        with retry_lock:
                            retry_results[r_idx] = ("fail", acct_label_r)
                        self._msg_queue.put(("error", f"✗ 재시도 실패: {acct_label_r} [로그인 실패] — {e}"))
                        _log_err_r("로그인", "로그인 실패", str(e))
                    except InvalidParameter as e:
                        with retry_lock:
                            retry_results[r_idx] = ("fail", acct_label_r)
                        self._msg_queue.put(("error", f"✗ 재시도 실패: {acct_label_r} [코드 무효] — {e}"))
                        _log_err_r("코드 입력", "코드 무효", str(e))
                    except Exception as e:
                        stage_r, error_type_r = classify_exception(e)
                        with retry_lock:
                            retry_results[r_idx] = ("fail", acct_label_r)
                        self._msg_queue.put(("error", f"✗ 재시도 실패: {acct_label_r} [{error_type_r}] — {str(e)[:100]}"))
                        _log_err_r(stage_r, error_type_r, str(e))

                with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as pool:
                    futures = []
                    for r_idx, orig_i in enumerate(failed_indices):
                        futures.append(pool.submit(_retry_process, r_idx, orig_i))
                        if r_idx < len(failed_indices) - 1:
                            import random
                            delay = random.uniform(*STAGGER_DELAY_API) if API_OK else STAGGER_DELAY
                            if self._auto_cancel_event.wait(delay):
                                break
                    concurrent.futures.wait(futures)

                retry_cancelled = any(v[0] == "cancel" for v in retry_results.values())
                if retry_cancelled:
                    self._msg_queue.put(("system", "── 자동 입력 취소됨 ──"))
                    self._code_detected_this_slot = True

            success_count = sum(1 for v in results.values() if v[0] == "ok")
            fail_count    = sum(1 for v in results.values() if v[0] in ("fail", "login_fail"))
            skip_count    = total - len(results)
            self._success_count += success_count
            self._fail_count += fail_count
            self._last_code_for_status = code
            self._last_code_time_for_status = datetime.datetime.now().isoformat()
            tag = "auto_ok" if fail_count == 0 and skip_count == 0 else "system"
            summary = f"━━ 전체 완료: {code} | 성공 {success_count} / 실패 {fail_count}"
            if skip_count > 0:
                summary += f" / 미실행 {skip_count}"
            summary += f" / 총 {total} | {_dt.datetime.now().strftime('%H:%M:%S')} ━━"
            self._msg_queue.put((tag, summary))
        finally:
            # GUI에 자동입력 종료 알림
            self._msg_queue.put(("__auto_end__", ""))

    # ------------------------------------------------------------------
    # 로그 출력
    # ------------------------------------------------------------------

    def _append_log(self, text: str, tag: str = "") -> None:
        self._log_lines.append({"text": text, "tag": tag})
        if len(self._log_lines) > 10000:
            self._log_lines = self._log_lines[-10000:]
        self._log.config(state=tk.NORMAL)
        self._log.insert(tk.END, text + "\n", tag if tag else ())
        # Text 위젯 메모리 관리: 5000줄 초과 시 오래된 줄 삭제
        line_count = int(self._log.index("end-1c").split(".")[0])
        if line_count > 5000:
            self._log.delete("1.0", f"{line_count - 5000}.0")
        self._log.see(tk.END)
        self._log.config(state=tk.DISABLED)

    # ------------------------------------------------------------------
    # 로그 영구 저장 / 복원 / 내보내기
    # ------------------------------------------------------------------

    def _load_log_data(self) -> None:
        """프로그램 시작 시 log_data.json에서 이전 로그와 캐치 코드 복원."""
        if not LOG_DATA_PATH.exists():
            return
        try:
            with open(LOG_DATA_PATH, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return

        # 로그 라인 복원
        lines = data.get("log_lines", [])
        self._log_lines = lines[-2000:]
        if self._log_lines:
            self._log.config(state=tk.NORMAL)
            for item in self._log_lines:
                tag = item.get("tag", "")
                self._log.insert(tk.END, item.get("text", "") + "\n",
                                 tag if tag else ())
            self._log.see(tk.END)
            self._log.config(state=tk.DISABLED)

        # 캐치 코드 복원
        caught = data.get("caught_codes", [])
        self._caught_codes_full = caught
        if caught:
            latest = caught[-1]
            self._latest_code_var.set(latest["code"])
            self._latest_meta_var.set(f"{latest['ts']}  |  {latest['sender']}")
            prev = [c["code"] for c in reversed(caught[:-1])][:8]
            self._caught_history = prev
            self._history_var.set("  ".join(prev) if prev else "없음")
            self._status_var.set(
                f"이전 코드 복원: {latest['code']}  ({latest.get('datetime','')[:10]})")

    def _save_log_data(self) -> None:
        """종료 시 log_data.json에 로그·캐치 코드 저장 (원자적)."""
        try:
            data = {
                "caught_codes": self._caught_codes_full,
                "log_lines":    self._log_lines[-10000:],
            }
            tmp_path = LOG_DATA_PATH.with_suffix(".tmp")
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            tmp_path.replace(LOG_DATA_PATH)
        except Exception as e:
            logger.warning(f"로그 데이터 저장 실패: {e}")

    def _export_text(self) -> None:
        """캐치 코드 + 메시지 로그를 텍스트 파일로 내보내기."""
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")],
            initialfile=f"카카오모니터_{now_str}.txt",
            title="텍스트 파일로 저장",
        )
        if not path:
            return

        sep = "═" * 60
        out = []

        # ── 캐치된 코드 목록
        out.append(sep)
        out.append(f"  캐치된 코드 목록  (총 {len(self._caught_codes_full)}개)")
        out.append(sep)
        if self._caught_codes_full:
            for i, c in enumerate(self._caught_codes_full, 1):
                dt = c.get("datetime", "")[:19].replace("T", " ")
                out.append(f"  {i:>3}.  {c['code']}   {c['ts']}   {c['sender']}   {dt}")
        else:
            out.append("  (없음)")
        out.append("")

        # ── 메시지 로그
        out.append(sep)
        out.append("  메시지 로그")
        out.append(sep)
        for item in self._log_lines:
            out.append(item.get("text", ""))

        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(out))
            self._append_log(f"✓ 저장 완료: {path}", tag="auto_ok")
        except Exception as e:
            messagebox.showerror("저장 실패", str(e), parent=self)


# ---------------------------------------------------------------------------
# 계정 관리 다이얼로그
# ---------------------------------------------------------------------------

class _AccountEditDialog(tk.Toplevel):
    def __init__(self, parent, account: dict = None) -> None:
        super().__init__(parent)
        self.title("계정 추가" if account is None else "계정 수정")
        self.resizable(False, False)
        self.configure(bg=C["bg"])
        self.grab_set()
        self.result: Optional[dict] = None
        self.columnconfigure(1, weight=1)

        fields = [
            ("이메일",   "email",    False),
            ("비밀번호", "password", True),
            ("비고",     "memo",     False),
            ("프록시",   "proxy",    False),
        ]
        self._vars = {}
        for r, (label, key, is_pw) in enumerate(fields):
            ttk.Label(self, text=f"{label}:", style="TLabel").grid(
                row=r, column=0, padx=14, pady=6, sticky=tk.W)
            var = tk.StringVar(value=account.get(key, "") if account else "")
            self._vars[key] = var
            entry = ttk.Entry(self, textvariable=var, width=34,
                              show="*" if is_pw else "")
            entry.grid(row=r, column=1, padx=(0, 14), pady=6, sticky=tk.EW)
            if is_pw:
                self._pw_entry = entry

        # 비밀번호 표시 체크
        self._show_pw = tk.BooleanVar(value=False)
        ttk.Checkbutton(self, text="비밀번호 표시", variable=self._show_pw,
                        command=lambda: self._pw_entry.config(
                            show="" if self._show_pw.get() else "*")).grid(
            row=1, column=2, padx=(0, 14))

        # 프록시 힌트
        ttk.Label(self, text="예) http://1.2.3.4:8080  또는  http://user:pw@ip:port  (빈칸=미사용)",
                  foreground=C["fg_dim"], font=("Segoe UI", 7)).grid(
            row=3, column=1, sticky=tk.W, padx=(0, 14), pady=(0, 4))

        # 사용 여부
        self._enabled_var = tk.BooleanVar(
            value=account.get("enabled", True) if account else True)
        ttk.Checkbutton(self, text="이 계정 사용",
                        variable=self._enabled_var).grid(
            row=4, column=1, sticky=tk.W, padx=(0, 14), pady=(0, 6))

        btn_frame = ttk.Frame(self)
        btn_frame.grid(row=5, column=0, columnspan=3, pady=12)
        ttk.Button(btn_frame, text="확인", style="Start.TButton",
                   command=self._ok).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frame, text="취소",
                   command=self.destroy).pack(side=tk.LEFT, padx=6)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _ok(self) -> None:
        email = self._vars["email"].get().strip()
        if not email:
            messagebox.showwarning("입력 오류", "이메일을 입력하세요.", parent=self)
            return
        self.result = {
            "email":    email,
            "password": self._vars["password"].get(),
            "memo":     self._vars["memo"].get().strip(),
            "proxy":    self._vars["proxy"].get().strip(),
            "enabled":  self._enabled_var.get(),
        }
        self.destroy()


class ProxyManagerDialog(tk.Toplevel):
    def __init__(self, parent, proxies: list, on_save) -> None:
        super().__init__(parent)
        self.title("프록시 IP 관리")
        self.resizable(True, True)
        self.minsize(480, 400)
        self.configure(bg=C["bg"])
        self.grab_set()
        self._proxies: list = list(proxies)
        self._on_save = on_save
        self._build_ui()
        self._refresh()

    def _build_ui(self) -> None:
        ttk.Label(self,
                  text="  워커 순서대로 자동 배분  |  형식: http://ip:port  또는  http://user:pw@ip:port",
                  foreground=C["fg_dim"], font=("Segoe UI", 8)).pack(
            padx=10, pady=(8, 2), anchor=tk.W)

        list_frame = tk.Frame(self, bg=C["bg"])
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        sb = ttk.Scrollbar(list_frame)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._listbox = tk.Listbox(
            list_frame, yscrollcommand=sb.set,
            bg=C["input"], fg=C["fg"], selectbackground=C["sel"],
            selectforeground=C["fg_bright"], font=("Consolas", 10),
            activestyle="none", relief=tk.FLAT, bd=0,
        )
        sb.config(command=self._listbox.yview)
        self._listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 입력 행
        entry_frame = tk.Frame(self, bg=C["bg"])
        entry_frame.pack(fill=tk.X, padx=10, pady=(0, 4))
        self._entry_var = tk.StringVar()
        tk.Entry(entry_frame, textvariable=self._entry_var,
                 bg=C["input"], fg=C["fg"], insertbackground=C["fg"],
                 relief=tk.FLAT, font=("Consolas", 10)).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(entry_frame, text="추가",
                   command=self._add_one).pack(side=tk.LEFT)
        self.bind("<Return>", lambda e: self._add_one())

        # 버튼 행
        btn_frame = ttk.Frame(self)
        btn_frame.pack(padx=10, pady=(0, 10))
        ttk.Button(btn_frame, text="선택 삭제",
                   command=self._delete_selected).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="전체 삭제",
                   command=self._clear_all).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="파일 불러오기",
                   command=self._import_file).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="저장 후 닫기", style="Start.TButton",
                   command=self._save_close).pack(side=tk.LEFT, padx=3)

    def _refresh(self) -> None:
        self._listbox.delete(0, tk.END)
        for i, p in enumerate(self._proxies):
            self._listbox.insert(tk.END, f"  {i+1:>3}.  {p}")

    def _add_one(self) -> None:
        val = self._entry_var.get().strip()
        if not val:
            return
        for line in val.splitlines():
            line = line.strip()
            if line and line not in self._proxies:
                self._proxies.append(line)
        self._entry_var.set("")
        self._refresh()

    def _delete_selected(self) -> None:
        sel = list(self._listbox.curselection())
        for i in reversed(sel):
            del self._proxies[i]
        self._refresh()

    def _clear_all(self) -> None:
        if not messagebox.askyesno("전체 삭제", "등록된 IP를 모두 삭제하시겠습니까?", parent=self):
            return
        self._proxies.clear()
        self._refresh()

    def _import_file(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("텍스트 파일", "*.txt"), ("모든 파일", "*.*")],
            title="IP 목록 파일 선택 (한 줄에 하나씩)",
            parent=self,
        )
        if not path:
            return
        try:
            with open(path, encoding="utf-8") as f:
                lines = [l.strip() for l in f if l.strip()]
        except Exception as e:
            messagebox.showerror("불러오기 실패", str(e), parent=self)
            return
        added = 0
        for line in lines:
            if line not in self._proxies:
                self._proxies.append(line)
                added += 1
        self._refresh()
        messagebox.showinfo("완료", f"IP {added}개를 불러왔습니다.", parent=self)

    def _save_close(self) -> None:
        self._on_save(self._proxies)
        self.destroy()


class AccountManagerDialog(tk.Toplevel):
    def __init__(self, parent, accounts: list, current_idx: int, on_save) -> None:
        super().__init__(parent)
        self.title("계정 목록 관리")
        self.resizable(True, True)
        self.minsize(700, 460)
        self.configure(bg=C["bg"])
        self.grab_set()

        self._accounts: list = list(accounts)
        self._current_idx: int = current_idx
        self._on_save = on_save

        self._build_ui()
        self._refresh_list()

    def _build_ui(self) -> None:
        ttk.Label(self,
                  text="  ☑ 클릭 → 사용/미사용 전환    ✕ 클릭 → 삭제    더블클릭 → 수정",
                  foreground=C["fg_dim"],
                  font=("Segoe UI", 8)).pack(padx=10, pady=(8, 2), anchor=tk.W)
        self._all_enabled_var = tk.BooleanVar(value=True)

        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        cols = ("check", "no", "email", "password", "memo", "proxy", "del")
        self._tree = ttk.Treeview(tree_frame, columns=cols,
                                  show="headings", height=16)
        self._tree.heading("check",    text="사용")
        self._tree.heading("no",       text="No.")
        self._tree.heading("email",    text="이메일")
        self._tree.heading("password", text="비밀번호")
        self._tree.heading("memo",     text="비고")
        self._tree.heading("proxy",    text="프록시")
        self._tree.heading("del",      text="삭제")
        self._tree.column("check",    width=45,  anchor="center", stretch=False)
        self._tree.column("no",       width=45,  anchor="center", stretch=False)
        self._tree.column("email",    width=190)
        self._tree.column("password", width=70,  stretch=False)
        self._tree.column("memo",     width=90)
        self._tree.column("proxy",    width=140)
        self._tree.column("del",      width=45,  anchor="center", stretch=False)

        sb = ttk.Scrollbar(tree_frame, command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._tree.bind("<Double-1>",      self._on_double_click)
        self._tree.bind("<ButtonRelease-1>", self._on_tree_click)
        self._tree.tag_configure("current",  foreground=C["accent"])
        self._tree.tag_configure("disabled", foreground="#555555")

        btn_frame = ttk.Frame(self)
        btn_frame.pack(padx=10, pady=(4, 10))

        ttk.Button(btn_frame, text="추가",
                   command=self._add).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="수정",
                   command=self._edit).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Checkbutton(btn_frame, text="전체 사용",
                        variable=self._all_enabled_var,
                        command=self._toggle_all_enabled).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="엑셀 불러오기",
                   command=self._import_excel).pack(side=tk.LEFT, padx=3)
        ttk.Button(btn_frame, text="엑셀 저장",
                   command=self._export_excel).pack(side=tk.LEFT, padx=3)
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(
            side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_frame, text="저장 후 닫기", style="Start.TButton",
                   command=self._save_close).pack(side=tk.LEFT, padx=3)

    def _refresh_list(self) -> None:
        for item in self._tree.get_children():
            self._tree.delete(item)
        for i, acct in enumerate(self._accounts):
            enabled = acct.get("enabled", True)
            check  = "☑" if enabled else "☐"
            marker = "▶" if i == self._current_idx else str(i + 1)
            masked = "•" * min(len(acct.get("password", "")), 8) or "(없음)"
            memo   = acct.get("memo", "")
            proxy  = acct.get("proxy", "")
            proxy_disp = proxy if proxy else "—"
            tags = []
            if i == self._current_idx:
                tags.append("current")
            if not enabled:
                tags.append("disabled")
            self._tree.insert("", tk.END, iid=str(i),
                              values=(check, marker, acct.get("email", ""),
                                      masked, memo, proxy_disp, "✕"),
                              tags=tuple(tags))
        if self._accounts and 0 <= self._current_idx < len(self._accounts):
            self._tree.selection_set(str(self._current_idx))
            self._tree.see(str(self._current_idx))

        # 전체 선택 체크박스 동기화
        if self._accounts:
            all_on = all(a.get("enabled", True) for a in self._accounts)
            self._all_enabled_var.set(all_on)
        else:
            self._all_enabled_var.set(True)

    def _on_tree_click(self, event) -> None:
        if self._tree.identify_region(event.x, event.y) != "cell":
            return
        col    = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            return
        idx = int(row_id)
        if col == "#1":
            self._toggle_enabled(idx)
        elif col == "#7":
            self._delete_by_idx(idx)

    def _on_double_click(self, event) -> None:
        if self._tree.identify_column(event.x) in ("#1", "#7"):
            return
        self._edit()

    def _toggle_enabled(self, idx: int) -> None:
        self._accounts[idx]["enabled"] = not self._accounts[idx].get("enabled", True)
        self._refresh_list()

    def _delete_by_idx(self, idx: int) -> None:
        if not messagebox.askyesno(
                "삭제 확인",
                f"계정 {idx+1} ({self._accounts[idx]['email']})을 삭제하시겠습니까?",
                parent=self):
            return
        self._accounts.pop(idx)
        if self._current_idx >= len(self._accounts):
            self._current_idx = max(0, len(self._accounts) - 1)
        self._refresh_list()

    def _toggle_all_enabled(self) -> None:
        state = self._all_enabled_var.get()
        for a in self._accounts:
            a["enabled"] = state
        self._refresh_list()

    def _enable_all(self) -> None:
        for a in self._accounts:
            a["enabled"] = True
        self._refresh_list()

    def _disable_all(self) -> None:
        for a in self._accounts:
            a["enabled"] = False
        self._refresh_list()

    def _selected_idx(self) -> Optional[int]:
        sel = self._tree.selection()
        return int(sel[0]) if sel else None

    def _add(self) -> None:
        if len(self._accounts) >= 200:
            messagebox.showwarning("최대 200개", "계정은 최대 200개까지 등록 가능합니다.", parent=self)
            return
        dlg = _AccountEditDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            self._accounts.append(dlg.result)
            self._refresh_list()

    def _edit(self) -> None:
        idx = self._selected_idx()
        if idx is None:
            return
        dlg = _AccountEditDialog(self, self._accounts[idx])
        self.wait_window(dlg)
        if dlg.result:
            self._accounts[idx] = dlg.result
            self._refresh_list()

    def _import_excel(self) -> None:
        path = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx *.xls"), ("모든 파일", "*.*")],
            title="계정 목록 엑셀 파일 선택",
            parent=self,
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            messagebox.showerror("불러오기 실패", str(e), parent=self)
            return

        if not rows:
            messagebox.showwarning("내용 없음", "엑셀 파일이 비어 있습니다.", parent=self)
            return

        # 헤더 행 스킵
        start = 0
        first = str(rows[0][0] or "").strip().lower()
        if first in ("이메일", "email", "e-mail"):
            start = 1

        new_accounts = []
        for row in rows[start:]:
            if not row or not row[0]:
                continue
            email = str(row[0]).strip()
            if not email:
                continue
            pw    = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            memo  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            enabled_raw = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else "true"
            enabled = enabled_raw not in ("false", "x", "✕", "0", "n", "no", "미사용")
            new_accounts.append({"email": email, "password": pw,
                                  "memo": memo, "enabled": enabled})

        if not new_accounts:
            messagebox.showwarning("계정 없음", "불러올 계정이 없습니다.", parent=self)
            return

        mode = messagebox.askyesnocancel(
            "불러오기 방식",
            f"엑셀에서 계정 {len(new_accounts)}개를 찾았습니다.\n\n"
            "[예] 기존 목록에 추가\n"
            "[아니오] 기존 목록을 대체\n"
            "[취소] 취소",
            parent=self,
        )
        if mode is None:
            return
        if mode:
            self._accounts.extend(new_accounts)
        else:
            self._accounts = new_accounts
            self._current_idx = 0
        self._refresh_list()
        messagebox.showinfo("완료", f"계정 {len(new_accounts)}개를 불러왔습니다.", parent=self)

    def _export_excel(self) -> None:
        if not self._accounts:
            messagebox.showwarning("계정 없음", "내보낼 계정이 없습니다.", parent=self)
            return
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
            initialfile=f"카카오모니터_계정_{now_str}.xlsx",
            title="계정 목록 저장",
            parent=self,
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "계정목록"

            headers = ["이메일", "비밀번호", "비고", "사용여부"]
            header_fill = PatternFill("solid", fgColor="2D6A2D")
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for r, acct in enumerate(self._accounts, 2):
                ws.cell(row=r, column=1, value=acct.get("email", ""))
                ws.cell(row=r, column=2, value=acct.get("password", ""))
                ws.cell(row=r, column=3, value=acct.get("memo", ""))
                ws.cell(row=r, column=4, value="사용" if acct.get("enabled", True) else "미사용")

            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 10
            wb.save(path)
        except Exception as e:
            messagebox.showerror("저장 실패", str(e), parent=self)
            return
        messagebox.showinfo("완료", f"계정 {len(self._accounts)}개를 저장했습니다.\n{path}", parent=self)

    def _save_close(self) -> None:
        self._on_save(self._accounts, self._current_idx)
        self.destroy()


# ---------------------------------------------------------------------------
# 진입점
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = App()
    app._win.mainloop()
